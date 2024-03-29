import types
import openpyxl
from io import BytesIO
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.read_only import EMPTY_CELL
from django.conf import settings
from django.db.models import QuerySet
from accounts.models import ProgramStudi
from learning_outcomes_assessment.utils import _iter_cols
from .models import (
    MataKuliahSemester,
    PesertaMataKuliah,
    NilaiExcelMataKuliahSemester,
)
from clo.models import KomponenClo, NilaiKomponenCloPeserta


def process_excel_file(
    mk_semester: MataKuliahSemester, 
    list_komponen_clo: QuerySet[KomponenClo]
):
    is_import_success = False
    message = ''
    import_result = {}

    # Get file
    try:
        nilai_excel_obj = NilaiExcelMataKuliahSemester.objects.get(
            mk_semester=mk_semester
        )
    except (NilaiExcelMataKuliahSemester.DoesNotExist, NilaiExcelMataKuliahSemester.MultipleObjectsReturned):
        if settings.DEBUG:
            print('Nilai Excel object for MK Semester (id={}) not found'.format(mk_semester.pk))
        message = 'File excel tidak dapat ditemukan.'
        return (is_import_success, message, import_result)
    
    with open(nilai_excel_obj.file.path, 'rb') as excel_file:
        excel_in_memory_file = BytesIO(excel_file.read())
    
    workbook: Workbook = openpyxl.load_workbook(excel_in_memory_file, read_only=True)
    worksheet: Worksheet = workbook.active
    worksheet.iter_cols = types.MethodType(_iter_cols, worksheet)
    
    # Validate MK Semester
    unique_id: str = worksheet['C9'].value
    semester_prodi_id, mk_semester_id = unique_id.split('/')
    
    if mk_semester_id != str(mk_semester.pk) or semester_prodi_id != str(mk_semester.semester.id_neosia):
        message = 'Unique ID tidak sesuai. Ekspektasi: {}/{}, Excel: {}/{}'.format(mk_semester.semester.id_neosia, mk_semester.pk, semester_prodi_id, mk_semester_id)
        if settings.DEBUG: print(message)
        return (is_import_success, message, import_result)

    # Validate Komponen CLO
    clo_row = 16
    komponen_clo_row = 17
    persentase_komponen_row = 18

    list_clo_excel = {}
    current_clo = ''
    
    # Excel values
    for cell in worksheet[clo_row][3:]:
        if cell is EMPTY_CELL: continue

        komponen = worksheet['{}{}'.format(cell.column_letter, komponen_clo_row)].value
        persentase = worksheet['{}{}'.format(cell.column_letter, persentase_komponen_row)].value
        
        # Add new komponen in current CLO
        if cell.value is None:
            list_clo_excel[current_clo].append({
                'komponen': komponen,
                'persentase': persentase
            })
            continue
        
        # Add new CLO and its komponen
        current_clo = cell.value
        list_clo_excel[current_clo] = []
        
        if komponen is None or persentase is None: continue

        list_clo_excel[current_clo].append({
            'komponen': komponen,
            'persentase': persentase
        })

    komponen_index = 0

    # Check excel values
    for clo_excel in list_clo_excel.keys():
        list_komponen_clo_excel = list_clo_excel[clo_excel]

        # Check CLO order
        current_clo = list_komponen_clo[komponen_index].clo.nama
        if current_clo != clo_excel:
            message = 'Urutan CPMK tidak sesuai. Ekspektasi: {}, file excel:{}'.format(
                current_clo, clo_excel)
            if settings.DEBUG: print(message)
            return (is_import_success, message, import_result)
            
        komponen_qs = list_komponen_clo.filter(
            clo__nama=clo_excel,
        )

        # Check length
        if len(list_komponen_clo_excel) != komponen_qs.count():
            message = 'Banyak komponen "{}" tidak sama. Ekspektasi: {}, file excel: {}'.format(clo_excel, komponen_qs.count(), len(list_komponen_clo_excel))
            if settings.DEBUG: print(message)
            return (is_import_success, message, import_result)

        # Check component and Persentase
        for komponen_clo_excel in list_komponen_clo_excel:
            nama_komponen = komponen_clo_excel['komponen']
            persentase_komponen = komponen_clo_excel['persentase']
            
            # Check komponen order
            current_instrumen_penilaian = list_komponen_clo[komponen_index].instrumen_penilaian
            current_persentase_komponen = list_komponen_clo[komponen_index].persentase / 100

            if current_instrumen_penilaian != nama_komponen:
                message = 'Urutan komponen tidak sesuai. Ekspektasi: {}, file excel: {}'.format(current_instrumen_penilaian, nama_komponen)
                if settings.DEBUG: print(message)
                return (is_import_success, message, import_result)
            
            if current_persentase_komponen != persentase_komponen:
                message = 'Urutan persentase komponen tidak sesuai. Ekspektasi: {}, file excel: {}'.format(current_persentase_komponen, persentase_komponen)
                if settings.DEBUG: print(message)
                return (is_import_success, message, import_result)

            qs = list_komponen_clo.filter(
                clo__nama=clo_excel,
                instrumen_penilaian=nama_komponen,
                persentase=round(persentase_komponen*100, 10)
            )

            if not qs.exists():
                message = 'CPMK, komponen, dan persentasenya tidak sesuai (case sensitive). File excel: CPMK: {}, Komponen CPMK: {}, Persentase: {}'.format(clo_excel, nama_komponen, persentase_komponen)
                if settings.DEBUG: print(message)
                return (is_import_success, message, import_result)
            
            komponen_index += 1
 
    # Validate peserta
    list_peserta_mk = PesertaMataKuliah.objects.filter(
        kelas_mk_semester__mk_semester=mk_semester
    ).order_by('kelas_mk_semester__nama', 'mahasiswa__username')
    mahasiswa_row = 19

    # Check mahasiswa length
    # 18 is row before 19 (start row of list mahasiswa)
    len_list_mahasiswa_excel = len(worksheet['B']) - 18
    if len_list_mahasiswa_excel != list_peserta_mk.count():
        message = 'Banyak mahasiswa tidak sesuai. Ekspektasi: {}, Excel: {}'.format(list_peserta_mk.count(), len_list_mahasiswa_excel)
        if settings.DEBUG: print(message)
        return (is_import_success, message, import_result)
    
    # Check mahasiswa
    for _ in range(len_list_mahasiswa_excel):
        nim = worksheet['B{}'.format(mahasiswa_row)].value
        nama = worksheet['C{}'.format(mahasiswa_row)].value

        qs = list_peserta_mk.filter(
            mahasiswa__username=nim,
            mahasiswa__nama=nama
        )

        if not qs.exists():
            message = 'Data mahasiswa tidak ada di database. NIM: {}, Nama: {}'.format(nim, nama)
            if settings.DEBUG: print(message)
            return (is_import_success, message, import_result)
        
        # Add nilai to import result
        if nim not in import_result.keys():
            import_result[nim] = [cell.value for cell in worksheet[mahasiswa_row][3:]]

        mahasiswa_row += 1

    # Check import result
    if len(import_result.keys()) == len_list_mahasiswa_excel:
        is_import_success = True
        message = 'Berhasil meng-import nilai dari file Excel.'
    else:
        message = 'Banyak mahasiswa yang berhasil diimport tidak sesuai dengan banyak mahasiswa di database. Ekspektasi: {}, yang berhasil: {}'.format(len_list_mahasiswa_excel, len(import_result.keys()))
    
    if settings.DEBUG: print(message)
    return (is_import_success, message, import_result)


def nilai_komponen_edit_process(cleaned_data: dict, prodi: ProgramStudi):
    is_success = False
    list_nilai_akhir_peserta = {}

    for nilai_komponen_clo_submit in cleaned_data:
        # If dict is empty, skip
        if not nilai_komponen_clo_submit: continue

        peserta: PesertaMataKuliah = nilai_komponen_clo_submit.get('peserta')
        komponen_clo: KomponenClo = nilai_komponen_clo_submit.get('komponen_clo')
        nilai_komponen = nilai_komponen_clo_submit.get('nilai')

        # Check query first
        nilai_peserta_qs = NilaiKomponenCloPeserta.objects.filter(
            peserta=peserta,
            komponen_clo=komponen_clo
        )

        # If exists, then update the query
        if nilai_peserta_qs.exists():
            # Use this save() method to trigger pre_save and post_save
            nilai_peserta_obj = nilai_peserta_qs.first()
            nilai_peserta_obj.nilai = nilai_komponen
            nilai_peserta_obj.save()
        else:  # Else, create new one
            # If form is empty, skip
            if len(nilai_komponen_clo_submit.items()) == 0: continue

            # Create new one
            NilaiKomponenCloPeserta.objects.create(**nilai_komponen_clo_submit)

        if peserta is not None:
            # Set nim peserta to list nilai akhir
            nim = peserta.mahasiswa.username
            if nim not in list_nilai_akhir_peserta.keys():
                list_nilai_akhir_peserta[nim] = {
                    'peserta_id': peserta.id_neosia,
                    'nilai_akhir': 0
                }
            
            # Set nilai
            if nilai_komponen is not None:
                list_nilai_akhir_peserta[nim]['nilai_akhir'] += (komponen_clo.persentase/100) * nilai_komponen

    if not prodi.is_restricted_mode:
        for nim, nilai_akhir_peserta in list_nilai_akhir_peserta.items():
            peserta: PesertaMataKuliah = PesertaMataKuliah.objects.get(id_neosia=nilai_akhir_peserta['peserta_id'])
            nilai_akhir = nilai_akhir_peserta['nilai_akhir']
            nilai_huruf = None

            if 85 <= nilai_akhir <= 100:
                nilai_huruf = 'A'
            elif 80 <= nilai_akhir < 85:
                nilai_huruf = 'A-'
            elif 75 <= nilai_akhir < 80:
                nilai_huruf = 'B+'
            elif 70 <= nilai_akhir < 75:
                nilai_huruf = 'B'
            elif 65 <= nilai_akhir < 70:
                nilai_huruf = 'B-'
            elif 60 <= nilai_akhir < 65:
                nilai_huruf = 'C+'
            elif 50 <= nilai_akhir < 60:
                nilai_huruf = 'C'
            elif 40 <= nilai_akhir < 50:
                nilai_huruf = 'D'
            elif nilai_akhir < 40:
                nilai_huruf = 'E'

            peserta.nilai_akhir = nilai_akhir
            peserta.nilai_huruf = nilai_huruf
            peserta.save()

    is_success = True
    return is_success
