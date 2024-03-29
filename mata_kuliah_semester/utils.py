import os
import numpy as np
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from copy import copy
from io import BytesIO
from functools import partial
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, Alignment, Border, Side,
    Protection,
    PatternFill
)
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    Paragraph, Spacer, Table, ListFlowable, TableStyle,
    Frame, PageTemplate, Image, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.conf import settings
from django.db.models import QuerySet
from learning_outcomes_assessment.chart.utils import radar_factory, save_chart
from learning_outcomes_assessment.utils import request_data_to_neosia
from .models import (
    KelasMataKuliahSemester,
    MataKuliahSemester,
    PesertaMataKuliah,
    DosenMataKuliah,
    NilaiMataKuliahIloMahasiswa,
)
from learning_outcomes_assessment.utils import export_pdf_header
from clo.models import (
    Clo,
    KomponenClo,
    NilaiKomponenCloPeserta,
    NilaiCloMataKuliahSemester,
    NilaiCloPeserta,
)
from pi_area.models import PerformanceIndicator
from ilo.models import Ilo
from mata_kuliah_kurikulum.models import MataKuliahKurikulum


PRODI_SEMESTER_URL = 'https://customapi.neosia.unhas.ac.id/getProdiSemester'
MATA_KULIAH_SEMESTER_URL = 'https://customapi.neosia.unhas.ac.id/getKelasBySemester'
PESERTA_MATA_KULIAH_URL = 'https://customapi.neosia.unhas.ac.id/getMahasiswaByKelas'
DOSEN_MATA_KULIAH_URL = 'https://customapi.neosia.unhas.ac.id/getDosenByKelas'
matplotlib.use('Agg')


def get_kelas_mk_semester(semester_prodi_id: int):
    list_kelas_mk_semester = []

    # Get MK semester
    parameters = {
        'id_prodi_semester': semester_prodi_id
    }
    json_response = request_data_to_neosia(MATA_KULIAH_SEMESTER_URL, parameters)
    if json_response is None: return list_kelas_mk_semester

    for mk_semester_per_kelas in json_response:
        mata_kuliah = {
            'id': mk_semester_per_kelas['id'],
            'id_mata_kuliah': mk_semester_per_kelas['id_mata_kuliah'],
            'nama': mk_semester_per_kelas['nama']
        }

        list_kelas_mk_semester.append(mata_kuliah)

    return list_kelas_mk_semester


def get_peserta_kelas_mk_semester(mk_semester: MataKuliahSemester):
    list_peserta = []
    list_kelas_mk_semester: QuerySet[KelasMataKuliahSemester] = mk_semester.get_kelas_mk_semester()
    prodi = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi

    for kelas_mk_semester in list_kelas_mk_semester:
        parameters = {
            'id_kelas': kelas_mk_semester.id_neosia
        }

        json_response = request_data_to_neosia(PESERTA_MATA_KULIAH_URL, parameters)
        if json_response is None: continue

        for peserta_data in json_response:
            if peserta_data['id'] is None: continue

            peserta = {
                'id_neosia': peserta_data['id'],
                'id_kelas_mk_semester': peserta_data['id_kelas_kuliah'],
                'mahasiswa': {
                    'username': peserta_data['nim'],
                    'nama': peserta_data['nama_mahasiswa'],
                },
                'nama': peserta_data['nama_mahasiswa'],
            }

            # If on resticted mode, take nilai akhir, else dont take it
            if prodi.is_restricted_mode:
                try:
                    nilai_akhir_response = float(peserta_data['nilai_akhir'])
                except ValueError: 
                    if settings.DEBUG:
                        print('Cannot convert {} to float.'.format(peserta_data['nilai_akhir']))
                    continue
                except TypeError:
                    # Nilai akhir from Neosia maybe null
                    nilai_akhir_response = peserta_data['nilai_akhir']

                peserta.update({
                    'nilai_akhir': nilai_akhir_response,
                    'nilai_huruf': peserta_data['nilai_huruf'],
                })
            else:
                peserta.update({
                    'nilai_akhir': None,
                    'nilai_huruf': None,
                })
            
            list_peserta.append(peserta)

    return list_peserta


def get_dosen_kelas_mk_semester(kelas_mk_semester_id: int):
    list_dosen = []
    parameters = {
        'id_kelas': kelas_mk_semester_id
    }

    json_response = request_data_to_neosia(DOSEN_MATA_KULIAH_URL, parameters)
    if json_response is None: return list_dosen

    for dosen_data in json_response:
        dosen = {
            'nip': dosen_data['nip'],
            'nama': dosen_data['nama'],
            'id_prodi': dosen_data['id_prodi']
        }

        list_dosen.append(dosen)

    return list_dosen


def get_kelas_mk_semester_choices(semester_prodi_id: int):
    """Get mata kuliah semester choices for choice field
    Returns only mata kuliah kurikulum because all classess in mata kuliah
    semester will be synchronized

    Args:
        semester_prodi_id (int): Semester Prodi ID

    Returns:
        list: List mata kuliah semester
    """
    list_kelas_mk_semester = get_kelas_mk_semester(semester_prodi_id)
    kelas_mk_semester_choices = []

    for mk_semester_per_kelas in list_kelas_mk_semester:
        id_mk_kurikulum = mk_semester_per_kelas['id_mata_kuliah']
        id_kelas_mk_semester = mk_semester_per_kelas['id']

        kelas_mk_semester_qs = KelasMataKuliahSemester.objects.filter(id_neosia=id_kelas_mk_semester)

        # If kelas MK semester is already in database, skip 
        if kelas_mk_semester_qs.exists(): continue

        # Check whether mata kuliah kurikulum exist in database
        try:
            mk_kurikulum_obj = MataKuliahKurikulum.objects.get(id_neosia=id_mk_kurikulum)
        except MataKuliahKurikulum.DoesNotExist or MataKuliahKurikulum.MultipleObjectsReturned:
            continue

        kelas_mk_semester = {
            'id_neosia': id_kelas_mk_semester,
            'kode': mk_kurikulum_obj.kode,
            'nama': mk_semester_per_kelas['nama'],
            'sks': mk_kurikulum_obj.sks,
        }

        kelas_mk_semester_choice = kelas_mk_semester['id_neosia'], kelas_mk_semester
        kelas_mk_semester_choices.append(kelas_mk_semester_choice)

    return kelas_mk_semester_choices


def get_update_kelas_mk_semester_choices(mk_semester: MataKuliahSemester):
    list_kelas_mk_semester: QuerySet[KelasMataKuliahSemester] = mk_semester.get_kelas_mk_semester()
    list_kelas_mk_semester_response = get_kelas_mk_semester(mk_semester.semester.id_neosia)
    update_kelas_mk_semester_choices = []

    for kelas_mk_semester in list_kelas_mk_semester:
        for kelas_mk_semester_response in list_kelas_mk_semester_response:
            if kelas_mk_semester.id_neosia != kelas_mk_semester_response['id']: continue

            try:
                mk_kurikulum_obj = MataKuliahKurikulum.objects.get(id_neosia=kelas_mk_semester_response['id_mata_kuliah'])
            except MataKuliahKurikulum.DoesNotExist or MataKuliahKurikulum.MultipleObjectsReturned:
                continue
            
            # Check:
            # *Kelas MK Semester > nama
            # *Kelas MK Semester > MK Semester > MK Kurikulum > id neosia
            isDataOkay = kelas_mk_semester.nama == kelas_mk_semester_response['nama']
            
            if isDataOkay: break

            # Just to show table strip in update choice field (list view model C)
            kelas_mk_semester_response.update({
                'mk_semester': {
                    'mk_kurikulum': mk_kurikulum_obj
                }
            })
            update_kelas_mk_semester_data = {
                'new': kelas_mk_semester_response,
                'old': kelas_mk_semester,
            }

            update_kelas_mk_semester_choice = kelas_mk_semester.id_neosia, update_kelas_mk_semester_data
            update_kelas_mk_semester_choices.append(update_kelas_mk_semester_choice)

            break
    
    return update_kelas_mk_semester_choices


def get_peserta_kelas_mk_semester_choices(mk_semester: MataKuliahSemester):
    peserta_choices = []

    list_peserta_response = get_peserta_kelas_mk_semester(mk_semester)

    for peserta in list_peserta_response:
        id_peserta = peserta['id_neosia']
        
        peserta_qs = PesertaMataKuliah.objects.filter(id_neosia=id_peserta)
        
        # Skip peserta, if already in database
        if peserta_qs.exists(): continue

        peserta_choice = id_peserta, peserta
        peserta_choices.append(peserta_choice)
    
    return peserta_choices


def get_update_peserta_mk_semester_choices(mk_semester: MataKuliahSemester):
    list_peserta_mk_semester: QuerySet[PesertaMataKuliah] = mk_semester.get_all_peserta_mk_semester()
    list_peserta_mk_semester_response = get_peserta_kelas_mk_semester(mk_semester)
    update_peserta_mk_semester_choices = []
    prodi = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi

    for peserta in list_peserta_mk_semester:
        for peserta_response in list_peserta_mk_semester_response:
            if peserta.id_neosia != peserta_response['id_neosia']: continue
            
            # If prodi is on restricted mode, check nilai, else no need
            if prodi.is_restricted_mode:
                # Check:
                # *Nilai akhir
                # *Nilai huruf
                isDataOkay = peserta.nilai_akhir == peserta_response['nilai_akhir'] and peserta.nilai_huruf == peserta_response['nilai_huruf']
            else:
                isDataOkay = True
            
            if isDataOkay: break

            update_peserta_mk_data = {
                'new': peserta_response,
                'old': peserta
            }

            update_peserta_mk_choice = peserta.id_neosia, update_peserta_mk_data

            update_peserta_mk_semester_choices.append(update_peserta_mk_choice)
            break
    
    return update_peserta_mk_semester_choices


def calculate_nilai_per_clo_mk_semester(mk_semester: MataKuliahSemester):
    list_clo: QuerySet[Clo] = mk_semester.get_all_clo()
    
    average_clo_achievement = 0

    # Loop through CLO
    for clo_obj in list_clo:
        list_komponen_clo: QuerySet[KomponenClo] = clo_obj.get_komponen_clo()
        list_assessment_form = np.zeros((1, len(list_komponen_clo)))
        list_persentase_komponen = np.zeros((len(list_komponen_clo), 1))

        # Loop through komponen CLO
        for i, komponen_clo_obj in enumerate(list_komponen_clo):
            nilai_komponen_all_peserta = NilaiKomponenCloPeserta.objects.filter(
                komponen_clo=komponen_clo_obj
            )

            # Calculate average of list nilai komponen of all peserta
            average_assessment_form = np.average([nilai_komponen_peserta.nilai for nilai_komponen_peserta in nilai_komponen_all_peserta])
            
            # If average assessment form is NaN, set it to 0
            if np.isnan(average_assessment_form):
                average_assessment_form = 0

            list_assessment_form[0][i] = average_assessment_form
            list_persentase_komponen[i][0] = komponen_clo_obj.persentase
        
        # Loop through average assessment form to get clo achievement
        clo_achievement = 0
        clo_percentage = clo_obj.get_total_persentase_komponen()
        # Calculate CLO achievement
        clo_achievement = 1/clo_percentage * (list_assessment_form @ list_persentase_komponen).flatten()[0]

        # Save clo_achievement to database
        nilai_clo_mk_semester, _ = NilaiCloMataKuliahSemester.objects.get_or_create(
            clo=clo_obj,
            mk_semester=mk_semester,
        )

        nilai_clo_mk_semester.nilai = clo_achievement
        nilai_clo_mk_semester.save()

        # Sum all CLO Achievement
        average_clo_achievement += clo_percentage/100 * clo_achievement
    
    if mk_semester.average_clo_achievement is None:
        mk_semester.average_clo_achievement = average_clo_achievement
        mk_semester.save()
    else:
        if mk_semester.average_clo_achievement != average_clo_achievement:
            mk_semester.average_clo_achievement = average_clo_achievement
            mk_semester.save()


def calculate_nilai_per_clo_peserta(peserta: PesertaMataKuliah):
    list_clo: QuerySet[Clo] = peserta.kelas_mk_semester.mk_semester.get_all_clo()

    # Loop through CLO
    for clo_obj in list_clo:
        list_komponen_clo: QuerySet[KomponenClo] = clo_obj.get_komponen_clo()
        list_assessment_form = np.zeros((1, len(list_komponen_clo)))
        list_persentase_komponen = np.zeros((len(list_komponen_clo), 1))

        # Loop through komponen CLO
        for i, komponen_clo_obj in enumerate(list_komponen_clo):
            # Get nilai komponen CLO peserta
            nilai_komponen_peserta: QuerySet[NilaiKomponenCloPeserta] = peserta.get_nilai_komponen_clo_peserta(komponen_clo_obj)
            
            if nilai_komponen_peserta.first() is None: average_assessment_form = 0
            else:
                average_assessment_form = nilai_komponen_peserta.first().nilai

            list_assessment_form[0][i] = average_assessment_form
            list_persentase_komponen[i][0] = komponen_clo_obj.persentase
        
        # Loop through average assessment form to get clo achievement
        clo_achievement = 0
        clo_percentage = clo_obj.get_total_persentase_komponen()
        # Calculate CLO achievement
        clo_achievement = 1/clo_percentage * (list_assessment_form @ list_persentase_komponen).flatten()[0]

        # Save clo_achievement to database
        nilai_clo_peserta, _ = NilaiCloPeserta.objects.get_or_create(
            clo=clo_obj,
            peserta=peserta,
        )

        nilai_clo_peserta.nilai = clo_achievement
        nilai_clo_peserta.save()


def calculate_nilai_per_ilo_mahasiswa(peserta: PesertaMataKuliah):
    mk_semester = peserta.kelas_mk_semester.mk_semester
    list_ilo: QuerySet[Ilo] = Ilo.objects.filter(
        pi_area__performanceindicator__piclo__clo__mk_semester=mk_semester
    ).distinct()

    # Calculate all needed components
    max_sks_prodi = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.total_sks_lulus
    bobot_mk = mk_semester.mk_kurikulum.sks / max_sks_prodi
    nilai_max = 100

    # Loop through ILO
    for ilo in list_ilo:
        list_clo_ilo = Clo.objects.filter(
            mk_semester=mk_semester,
            piclo__performance_indicator__pi_area__ilo=ilo,
        ).distinct()

        # Get list persentase CLO
        list_persentase_clo = np.array([
            clo.get_total_persentase_komponen()/100 for clo in list_clo_ilo
        ])

        # Get list bobot PI
        list_pi_ilo = PerformanceIndicator.objects.filter(
            pi_area__ilo=ilo
        )
        list_bobot_pi_ilo = np.array([
            clo.get_pi_clo().count() / list_pi_ilo.count() for clo in list_clo_ilo
        ])

        # Get list nilai CLO peserta
        nilai_clo_peserta_qs = NilaiCloPeserta.objects.filter(
            peserta=peserta,
            clo__in=list_clo_ilo,
        ).values_list('nilai')
        list_nilai_clo_peserta = np.array(nilai_clo_peserta_qs).flatten()
        
        # Use try to prevent miss shape
        try:
            konversi_clo_to_ilo = bobot_mk * (list_persentase_clo * list_bobot_pi_ilo) * nilai_max
        except ValueError:
            if settings.DEBUG:
                print('Shape array tidak sama. Persentase CPMK: {}, Bobot PI CPL: {}, Nilai CPMK Peserta: {}'.format(
                    list_persentase_clo.shape, list_bobot_pi_ilo.shape, list_nilai_clo_peserta.shape))
            continue
        
        persentase_konversi = konversi_clo_to_ilo / np.sum(konversi_clo_to_ilo)
        
        # Use try to prevent miss shape
        try:
            persentase_capaian_ilo = persentase_konversi * list_nilai_clo_peserta
        except ValueError:
            if settings.DEBUG:
                print('Shape array tidak sama. Persentase Konversi: {}, Nilai CPMK Peserta: {}'.format(
                    persentase_konversi.shape, list_nilai_clo_peserta.shape))
            continue

        persentase_capaian_ilo = np.sum(persentase_capaian_ilo)

        nilai_ilo_mhs_obj, _ = NilaiMataKuliahIloMahasiswa.objects.get_or_create(
            peserta=peserta,
            ilo=ilo
        )

        if nilai_ilo_mhs_obj.nilai_ilo is None:
            nilai_ilo_mhs_obj.nilai_ilo = persentase_capaian_ilo
            nilai_ilo_mhs_obj.save()
        else:
            if nilai_ilo_mhs_obj.nilai_ilo != persentase_capaian_ilo:
                nilai_ilo_mhs_obj.nilai_ilo = persentase_capaian_ilo
                nilai_ilo_mhs_obj.save()


def alphabet_to_number(alphabet):
    """
    Change alphabet to number
    >>> alphabe_to_number('z')
    26
    """
    list_alphabet = list('abcdefghijklmnopqrstuvwxyz')
    return [i for i, huruf in enumerate(list_alphabet) if huruf == alphabet.lower()][0] + 1


def number_to_alphabet(number):
    """
    Change number to alphabet
    >>> number_to_alphabet(26)
    'z'
    """
    list_alphabet = list('abcdefghijklmnopqrstuvwxyz')
    return list_alphabet[number-1]


def add_border_to_merged_cells(merged_cells, border):
    """Add border to merged cells

    Args:
        merged_cells (_type_): Merged cells
        border (Border): Border to apply
    """
    
    for cells in merged_cells:
        for cell in cells:
            cell.border = border


def make_cell_title_style(cell):
    """Make cell style with title style, Boild and center alignment

    Args:
        cell (Cell): Target cell
    """

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    cell.font = bold_font
    cell.alignment = center_align


def stretch_cell_width(worksheet, column, column_letter):
    """Stretch cell width according to max width

    Args:
        worksheet (WorkSheet): Worksheet
        column (Column): Worksheet column
        column_letter (str): Column letter
    """

    max_width = 10

    for cell in column:
        if not cell.value: continue
        max_width = max((max_width, len(str(cell.value))))

    worksheet.column_dimensions[column_letter].width = max_width + 2


def lock_or_unlock_cells(cells_by_row, protection):
    """Lock or unlock cells

    Args:
        cells_by_row (list): Range Cells
        protection (Protection): Protection
    """
    
    for cells in cells_by_row:
        for cell in cells:
            cell.protection = protection


def fill_cells(cells_by_row, fill):
    """Fill cell with color

    Args:
        cells_by_row (list): Range Cells
        fill (PatternFill): Fill color
    """

    for cells in cells_by_row:
        for cell in cells:
            cell.fill = fill


def generate_template_nilai_mk_semester(mk_semester: MataKuliahSemester, list_peserta_mk: QuerySet[PesertaMataKuliah]):
    workbook = openpyxl.Workbook()
    worksheet: Worksheet  = workbook.active

    # Worksheet properties
    all_side_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    title = 'UNIVERSITAS HASANUDDIN'
    DETAIL = {
        'Fakultas': mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.fakultas.nama,
        'Program Studi': mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.nama,
        'Jenjang Prodi': mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.jenjang_studi.kode,
        'Semester': mk_semester.semester.semester.nama,
        'Kode Mata Kuliah': mk_semester.mk_kurikulum.kode,
        'Nama Mata Kuliah': mk_semester.mk_kurikulum.nama,
        'Unique ID': '{}/{}'.format(mk_semester.semester.pk, mk_semester.pk)
    }

    # Title
    worksheet.merge_cells('B1:H1')
    title_cell = worksheet['B1']
    
    title_cell.value = title
    make_cell_title_style(title_cell)

    # Height and width
    # 50 = 100px
    worksheet.row_dimensions[1].height = 50
    # 2 because it is what it is
    worksheet.column_dimensions['B'].width = len(title) + 2
    
    # Detail
    start_col = 'B'
    start_row_clo = 3
    for key, value in DETAIL.items():
        # Merge cells
        merged_cells = 'C{0}:H{0}'.format(start_row_clo)
        worksheet.merge_cells(merged_cells)

        # Column B
        cell = '{}{}'.format(start_col, start_row_clo)
        worksheet[cell] = key
        worksheet[cell].border = all_side_border

        # Column C
        cell = '{}{}'.format('C', start_row_clo)
        worksheet[cell] = value
        # Add border to merged cell
        add_border_to_merged_cells(worksheet[merged_cells], all_side_border)

        start_row_clo += 1
    
    # Catatan
    CATATAN = (
        '*User hanya bisa menambah, mengubah, dan menghapus nilai per Komponen dalam Sheet ini.',
        '*Disarankan untuk tidak mengubah "Unique ID" untuk meminimalisir kesalahan baca oleh sistem.',
        '*Disarankan untuk tidak membuka kunci Sheet untuk meminimalisir kesalahan baca oleh sistem.',
        '*Jika ingin menambah, menghapus, atau mengubah peserta, disarankan untuk mengubah dari sistemnya.',
    )
    start_col = 'B'
    end_col = 'H'
    start_col_num = alphabet_to_number(start_col)
    start_row = 11

    for i, note in enumerate(CATATAN, 1):
        merged_cells = '{0}{1}:{2}{1}'.format(start_col, start_row, end_col)
        worksheet.merge_cells(merged_cells)

        worksheet['{}{}'.format(start_col, start_row)] = note
        # +1 is next row
        start_row += 1

    # List mahasiswa dan nilai
    # Header
    HEADER_TITLES = ('No.', 'NIM', 'Nama')
    start_col = 'A'
    start_col_num = alphabet_to_number(start_col)
    start_row = 16
    end_row = 18

    for i, title in enumerate(HEADER_TITLES, 1):
        merged_cells = '{0}{1}:{0}{2}'.format(start_col, start_row, end_row)
        worksheet.merge_cells(merged_cells)

        # Title
        title_cell = '{}{}'.format(start_col, start_row)
        worksheet[title_cell] = title
        make_cell_title_style(worksheet[title_cell])
        add_border_to_merged_cells(worksheet[merged_cells], all_side_border)

        start_col = number_to_alphabet(start_col_num + i)

    # List CLO
    start_col = 'D'
    start_row_clo = 16
    start_row_komponen = 17
    start_row_persentase = 18

    list_clo: QuerySet[Clo] = mk_semester.get_all_clo()

    for clo in list_clo:
        start_col_num = alphabet_to_number(start_col)
        list_komponen: QuerySet[KomponenClo] = clo.get_komponen_clo().order_by('instrumen_penilaian')
        # Nama CLO
        # Merged cells
        if list_komponen.count() > 1:
            # -1 is start column itself
            end_col_num = start_col_num + len(list_komponen) - 1
            end_col = number_to_alphabet(end_col_num)
            
            merged_cells = '{0}{1}:{2}{1}'.format(start_col, start_row_clo, end_col)
            worksheet.merge_cells(merged_cells)
            add_border_to_merged_cells(worksheet[merged_cells], all_side_border)
        
        # Nama CLO
        cell = '{}{}'.format(start_col, start_row_clo)
        worksheet[cell] = clo.nama
        worksheet[cell].border = all_side_border
        make_cell_title_style(worksheet[cell])

        # Komponen CLO
        for komponen in list_komponen:
            column = number_to_alphabet(start_col_num)

            # Nama Komponen
            cell = '{}{}'.format(column, start_row_komponen)
            worksheet[cell] = komponen.instrumen_penilaian
            worksheet[cell].border = all_side_border
            make_cell_title_style(worksheet[cell])

            # Persentase Komponen
            cell = '{}{}'.format(column, start_row_persentase) 
            worksheet[cell] = komponen.persentase/100
            worksheet[cell].style = 'Percent'
            worksheet[cell].border = all_side_border
            make_cell_title_style(worksheet[cell])

            stretch_cell_width(worksheet, worksheet[column], column)

            # +1 is next column
            start_col_num += 1
        
        start_col = number_to_alphabet(start_col_num)

    # Set end column to add border
    end_col = column

    # List Mahasiswa
    start_row = 19

    for i, peserta_mk in enumerate(list_peserta_mk, 1):
        # Nomor
        cell = 'A{}'.format(start_row)
        worksheet[cell] = i
        worksheet[cell].alignment = center_align

        # NIM
        worksheet['B{}'.format(start_row)] = peserta_mk.mahasiswa.username
        # Nama
        worksheet['C{}'.format(start_row)] = peserta_mk.mahasiswa.nama

        start_row += 1

    end_row = start_row - 1
    add_border_to_merged_cells(worksheet['A19:{}{}'.format(end_col, end_row)], all_side_border)
    stretch_cell_width(worksheet, worksheet['C'], 'C')

    # Lock cells
    nilai_protection = Protection(locked=False, hidden=False)
    nilai_fill = PatternFill('solid', fgColor='fde047')
    nilai_cells = 'D19:{}{}'.format(end_col, end_row)
    lock_or_unlock_cells(worksheet[nilai_cells], nilai_protection)
    fill_cells(worksheet[nilai_cells], nilai_fill)
    worksheet.protection.enable()
    
    # Save the workbook to a BytesIO object
    file_stream = BytesIO()
    workbook.save(file_stream)

    file_stream.seek(0)

    return file_stream


def generate_nilai_file(mk_semester: MataKuliahSemester, list_nilai_huruf: dict):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    pdf_file = SimpleDocTemplate(file_stream)
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    font_size = normal_style.fontSize
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        mk_semester.mk_kurikulum.nama,
        style=styles['h1']
    )

    fakultas = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.fakultas.nama
    prodi = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.nama
    # MK Semester detail
    detail_data = [
        ['Fakultas', ':', fakultas],
        ['Program Studi', ':', prodi],
        ['Jenjang Studi', ':', mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.jenjang_studi.kode],
        ['Semester', ':', mk_semester.semester.semester.nama],
        ['Kode', ':', mk_semester.mk_kurikulum.kode],
        ['SKS', ':', mk_semester.mk_kurikulum.sks],
        ['Dosen', ':']
    ]
    detail_table_style = TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
    ])
    detail = Table(
        data=detail_data,
        style=detail_table_style,
        hAlign='LEFT'
    )
    list_dosen_mk_semester: list[DosenMataKuliah] = mk_semester.get_all_dosen_mk_semester()

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas, prodi), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])

    list_dosen = ListFlowable(
        [Paragraph('{}'.format(dosen_mk_semester.dosen.nama)) for dosen_mk_semester in list_dosen_mk_semester],
        start='1',
        bulletFontSize=font_size,
        bulletFormat='%s.'
    )

    # Nilai Mahasiswa
    mahasiswa_data = [
        ['No.', 'NIM', 'Nama', 'Nilai angka', 'Nilai huruf']
    ]
    list_mahasiswa_mk_semester: list[PesertaMataKuliah] = mk_semester.get_all_peserta_mk_semester()
    
    # (COL, ROW)
    table_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]
    mahasiswa_table_style_data = copy(table_style_data)
    mahasiswa_table_style_data += [
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('ALIGN', (3, 1), (-2, -1), 'RIGHT'),               # Nilai align
    ]

    for i, mahasiswa in enumerate(list_mahasiswa_mk_semester, 1):
        # Add mahasiswa nilai
        mahasiswa_data.append(
            [
                '{}.'.format(i), 
                mahasiswa.mahasiswa.username,
                Paragraph(mahasiswa.mahasiswa.nama),
                mahasiswa.nilai_akhir,
                mahasiswa.nilai_huruf
            ],
        )

        # Add merge cells for nilai komponen
        mahasiswa_table_style_data += [
            ('SPAN', (0, i*2), (-1, i*2)),
            ('ALIGN', (0, i*2), (0, i*2), 'LEFT'),
            ('LEFTPADDING', (0, i*2), (0, i*2), 32),
        ]

        # Nilai CLO Mahasiswa
        list_nilai_komponen: QuerySet[NilaiKomponenCloPeserta] = mahasiswa.get_all_nilai_komponen_clo_peserta()
        nilai_mahasiswa_table_data = []
        for nilai_komponen in list_nilai_komponen:
            nilai_mahasiswa_table_data.append(
                [
                    nilai_komponen.komponen_clo.clo.nama,
                    nilai_komponen.komponen_clo.instrumen_penilaian,
                    '{}%'.format(nilai_komponen.komponen_clo.persentase),
                    ':',
                    nilai_komponen.nilai
                ]
            )
        mahasiswa_data.append([
            Table(
                nilai_mahasiswa_table_data,
                hAlign='LEFT',
            )
        ])

    mahasiswa_table_style = TableStyle(mahasiswa_table_style_data)
    mahasiswa_table = Table(
        mahasiswa_data,
        style=mahasiswa_table_style,
        hAlign='LEFT',
        colWidths=[1*cm, 3*cm, 6.5*cm, None, None],
        repeatRows=1,
    )

    # Pencapaian per CPMK
    pencapaian_per_cpmk_title = Paragraph(
        'Pencapaian Capaian Pembelajaran Mata Kuliah',
        style=styles['h2']
    )

    # Pencapaian per CPMK table
    pencapaian_per_cpmk_table_data = [
        ('CPMK', 'CPL', 'Persentase CPMK', 'Skor CPMK', 'Hasil CPMK')
    ]

    list_nilai_clo_mk_semester: QuerySet[NilaiCloMataKuliahSemester] = mk_semester.get_nilai_clo_mk_semester()
    pencapaian_per_cpmk_chart_data = {}

    for nilai_clo in list_nilai_clo_mk_semester:
        clo = nilai_clo.clo
        persentase_clo = clo.get_total_persentase_komponen()
        hasil_clo = persentase_clo/100 * nilai_clo.nilai
        hasil_clo = float('{:.2f}'.format(hasil_clo))

        pencapaian_per_cpmk_table_data.append((
            clo.nama,
            clo.get_ilo().nama,
            '{}%'.format(persentase_clo),
            float('{:.2f}'.format(nilai_clo.nilai)),
            hasil_clo
        ))

        pencapaian_per_cpmk_chart_data[clo.nama] = hasil_clo
    
    # Avg CLO achievement
    if mk_semester.average_clo_achievement:
        pencapaian_per_cpmk_table_data.append((
            Paragraph(
                'Total Capaian Pembelajaran Mata Kuliah',
                style=ParagraphStyle(
                    name='table_capaian_cpmk_footer',
                    alignment=TA_RIGHT,
                    fontName='Helvetica-Bold',
                    fontSize=10
                )
            ),
            '', '', '',
            float('{:.2f}'.format(mk_semester.average_clo_achievement)),
        ))
    
    # Make table and table style
    pencapaian_per_cpmk_table_style_data = copy(table_style_data)
    pencapaian_per_cpmk_table_style_data += [
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),               # Nilai align
        ('SPAN', (0, -1), (-2, -1)),                        # Total merge
        ('FONTNAME', (-1, -1), (-1, -1), 'Helvetica-Bold'),    # Total font
    ]
    pencapaian_per_cpmk_table_style = TableStyle(pencapaian_per_cpmk_table_style_data)
    pencapaian_per_cpmk_table = Table(
        pencapaian_per_cpmk_table_data,
        style=pencapaian_per_cpmk_table_style,
        hAlign='LEFT',
        repeatRows=1,
    )

    # Pie chart capaian per CPMK
    pencapaian_per_cpmk_labels = pencapaian_per_cpmk_chart_data.keys()
    pencapaian_per_cpmk_sizes = pencapaian_per_cpmk_chart_data.values()

    fig, ax = plt.subplots(layout='constrained')

    ax.pie(
        pencapaian_per_cpmk_sizes, 
        labels=pencapaian_per_cpmk_labels, 
        autopct=lambda percentage: '{:.2f}'.format(
            percentage / 100 * mk_semester.average_clo_achievement)
    )
    ax.set_title('Pencapaian per CPMK')
    # Save chart
    pencapaian_per_cpmk_chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
    pencapaian_per_cpmk_chart_path = save_chart(pencapaian_per_cpmk_chart_dir, fig, width=3, height=3)

    pencapaian_per_cpmk_chart_image = Image(pencapaian_per_cpmk_chart_path)

    # Bar chart nilai mahasiswa
    nilai_mhs_labels = list_nilai_huruf.keys()
    nilai_mhs_sizes = list_nilai_huruf.values()

    fig, ax = plt.subplots(layout='constrained')
    ax.bar(nilai_mhs_labels, nilai_mhs_sizes)
    # Styling
    ax.set_title('Distribusi Nilai Mahasiswa')
    ax.grid(True, axis='y', which='major')
    ax.grid(True, axis='x', which='minor')
    minor_locator = mticker.AutoMinorLocator(2) # set the number of minor intervals per major interval
    ax.xaxis.set_minor_locator(minor_locator)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')
    
    # Save chart
    nilai_mhs_chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
    nilai_mhs_chart_path = save_chart(nilai_mhs_chart_dir, fig, width=4.5, height=2.5)

    nilai_mhs_chart_image = Image(nilai_mhs_chart_path)

    # Build
    pdf_file.build([
        title, detail,
        list_dosen, empty_line,
        mahasiswa_table, PageBreak(),
        pencapaian_per_cpmk_title,
        pencapaian_per_cpmk_table, PageBreak(),
        pencapaian_per_cpmk_chart_image, empty_line,
        nilai_mhs_chart_image,
    ])
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream


def generate_all_student_performance_file(mk_semester: MataKuliahSemester):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    pdf_file = SimpleDocTemplate(file_stream)
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    font_size = normal_style.fontSize
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        'Student Performance - {}'.format(mk_semester.mk_kurikulum.nama),
        style=styles['h1']
    )

    fakultas = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.fakultas.nama
    prodi = mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi.nama
    # MK Semester detail
    detail_data = [
        ['Fakultas', ':', fakultas],
        ['Program Studi', ':', prodi],
        ['Jenjang Studi', ':', mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.jenjang_studi.kode],
        ['Semester', ':', mk_semester.semester.semester.nama],
        ['Kode', ':', mk_semester.mk_kurikulum.kode],
        ['SKS', ':', mk_semester.mk_kurikulum.sks],
        ['Dosen', ':']
    ]
    detail_table_style = TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
    ])
    detail = Table(
        data=detail_data,
        style=detail_table_style,
        hAlign='LEFT'
    )
    list_dosen_mk_semester: list[DosenMataKuliah] = mk_semester.get_all_dosen_mk_semester()

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas, prodi), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])

    list_dosen = ListFlowable(
        [Paragraph('{}'.format(dosen_mk_semester.dosen.nama)) for dosen_mk_semester in list_dosen_mk_semester],
        start='1',
        bulletFontSize=font_size,
        bulletFormat='%s.'
    )

    # Nilai Mahasiswa
    mahasiswa_data = [
        ['No.', 'NIM', 'Nama', 'Nilai angka', 'Nilai huruf']
    ]
    list_mahasiswa_mk_semester: list[PesertaMataKuliah] = mk_semester.get_all_peserta_mk_semester()
    
    # (COL, ROW)
    table_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]
    mahasiswa_table_style_data = copy(table_style_data)
    mahasiswa_table_style_data += [
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('ALIGN', (3, 1), (-2, -1), 'RIGHT'),               # Nilai align
    ]

    for i, mahasiswa in enumerate(list_mahasiswa_mk_semester, 1):
        # Add mahasiswa nilai
        mahasiswa_data.append(
            [
                '{}.'.format(i), 
                mahasiswa.mahasiswa.username,
                Paragraph(mahasiswa.mahasiswa.nama),
                mahasiswa.nilai_akhir,
                mahasiswa.nilai_huruf
            ],
        )

        # Add merge cells for nilai komponen
        mahasiswa_table_style_data += [
            ('SPAN', (0, i*2), (-1, i*2)),
            ('ALIGN', (0, i*2), (0, i*2), 'LEFT'),
            ('LEFTPADDING', (0, i*2), (0, i*2), 32),
        ]

        # Nilai CLO Mahasiswa
        list_nilai_clo: QuerySet[NilaiCloPeserta] = NilaiCloPeserta.objects.filter(
            peserta=mahasiswa,
        ).order_by('clo')
        list_nilai_ilo: QuerySet[NilaiMataKuliahIloMahasiswa] = NilaiMataKuliahIloMahasiswa.objects.filter(
            peserta=mahasiswa,
        ).order_by('ilo')
        nilai_mahasiswa_table_data = []
        
        for nilai_clo in list_nilai_clo:
            nilai_mahasiswa_table_data.append(
                [
                    nilai_clo.clo.nama,
                    '{}%'.format(nilai_clo.clo.get_total_persentase_komponen()),
                    ':',
                    '{:.2f}'.format(nilai_clo.nilai)
                ]
            )

        nilai_mahasiswa_table_data.append([])

        for nilai_ilo in list_nilai_ilo:
            nilai_mahasiswa_table_data.append(
                [
                    nilai_ilo.ilo.nama,
                    '',
                    ':',
                    '{:.2f}'.format(nilai_ilo.nilai_ilo)
                ]
            )

        mahasiswa_data.append([
            Table(
                nilai_mahasiswa_table_data,
                hAlign='LEFT',
            )
        ])

    mahasiswa_table_style = TableStyle(mahasiswa_table_style_data)
    mahasiswa_table = Table(
        mahasiswa_data,
        style=mahasiswa_table_style,
        hAlign='LEFT',
        colWidths=[1*cm, 3*cm, 6.5*cm, None, None],
        repeatRows=1,
    )


    # Build
    pdf_file.build([
        title, detail,
        list_dosen, empty_line,
        mahasiswa_table, PageBreak(),
    ])
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream


def generate_student_performance_file(peserta_mk_semester: PesertaMataKuliah):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    pagesize = landscape(letter)
    pdf_file = SimpleDocTemplate(file_stream, pagesize=pagesize)
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    font_size = normal_style.fontSize
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        'Student Performance',
        style=styles['h1']
    )

    program_studi = peserta_mk_semester.kelas_mk_semester.mk_semester.mk_kurikulum.kurikulum.prodi_jenjang.program_studi
    fakultas = program_studi.fakultas.nama
    prodi = program_studi.nama

    # Header
    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas, prodi), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])
    
    # Detail
    # MK Semester detail
    mk_semester = peserta_mk_semester.kelas_mk_semester.mk_semester
    mk_kurikulum = mk_semester.mk_kurikulum
    detail_data = [
        ['Nama mahasiswa', ':', peserta_mk_semester.mahasiswa.nama],
        ['Mata kuliah', ':', mk_kurikulum.nama],
        ['Semester', ':', mk_semester.semester.semester.nama],
        ['Kode', ':', mk_kurikulum.kode],
        ['SKS', ':', mk_kurikulum.sks],
        ['Nilai akhir', ':', peserta_mk_semester.nilai_akhir]
    ]
    detail_table_style = TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
    ])
    detail = Table(
        data=detail_data,
        style=detail_table_style,
        hAlign='LEFT'
    )

    # Perolehan nilai CLO
    perolehan_nilai_clo_title = Paragraph(
        'Perolehan Nilai CPMK',
        style=styles['h2']
    )

    # (COL, ROW)
    table_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]

    perolehan_nilai_clo_table_data = [
        ('CPMK', 'Nilai CPMK')
    ]

    list_clo: QuerySet[Clo] = peserta_mk_semester.kelas_mk_semester.mk_semester.get_all_clo()
    list_nilai_clo_peserta: QuerySet[NilaiCloPeserta] = NilaiCloPeserta.objects.filter(
        clo__in=list_clo,
        peserta=peserta_mk_semester,
    ).order_by('clo')
    
    nilai_clo_peserta_chart_data = {}
    for nilai_clo_peserta in list_nilai_clo_peserta:
        nama_clo = nilai_clo_peserta.clo.nama
        nilai_clo = float('{:.2f}'.format(nilai_clo_peserta.nilai))

        nilai_clo_peserta_chart_data[nama_clo] = nilai_clo
        perolehan_nilai_clo_table_data.append((
            nama_clo,
            nilai_clo
        ))

    perolehan_nilai_clo_table = Table(
        perolehan_nilai_clo_table_data,
        style=table_style_data,
        hAlign='LEFT',
        repeatRows=1,
    )

    fig, ax = plt.subplots()
    ax.barh(
        list(nilai_clo_peserta_chart_data.keys()),
        list(nilai_clo_peserta_chart_data.values()),
    )

    # Styling
    ax.grid(True, axis='y', which='minor')
    ax.grid(True, axis='x', which='major')
    minor_locator = mticker.AutoMinorLocator(2) # set the number of minor intervals per major interval
    ax.yaxis.set_minor_locator(minor_locator)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')

    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_xlabel('Nilai')
    ax.set_ylabel('CPMK')
    ax.set_title('Grafik Perolehan Nilai CPMK')
    
    # Save chart
    perolehan_nilai_clo_chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
    perolehan_nilai_clo_chart_path = save_chart(perolehan_nilai_clo_chart_dir, fig)

    perolehan_nilai_clo_chart_image = Image(perolehan_nilai_clo_chart_path)

    # Perolehan Nilai CPL
    perolehan_nilai_ilo_title = Paragraph(
        'Perolehan Nilai CPL',
        style=styles['h2']
    )

    list_nilai_ilo_peserta: QuerySet[NilaiMataKuliahIloMahasiswa] = peserta_mk_semester.get_nilai_ilo()

    # Perolehan Nilai CPL table
    perolehan_nilai_ilo_table_data = [
        ('CPL', 'Satisfactory Level', 'Nilai CPL')
    ]

    perolehan_nilai_ilo_table_style_data = copy(table_style_data)
    list_nama_ilo = []
    perolehan_nilai_ilo_chart_data = {
        'Satisfactory level': [],
        'Nilai mahasiswa': [],
    }

    for i, nilai_ilo_peserta in enumerate(list_nilai_ilo_peserta):
        nilai_ilo = float('{:.2f}'.format(nilai_ilo_peserta.nilai_ilo))
        nama_ilo = nilai_ilo_peserta.ilo.nama
        satisfactory_level_ilo = nilai_ilo_peserta.ilo.satisfactory_level

        perolehan_nilai_ilo_table_data.append((
            nama_ilo,
            satisfactory_level_ilo,
            nilai_ilo
        ))

        list_nama_ilo.append(nama_ilo)
        perolehan_nilai_ilo_chart_data['Satisfactory level'].append(satisfactory_level_ilo)
        perolehan_nilai_ilo_chart_data['Nilai mahasiswa'].append(nilai_ilo)

        # Give red color
        if nilai_ilo < nilai_ilo_peserta.ilo.satisfactory_level:
            col = 2
            row = i + 1
            perolehan_nilai_ilo_table_style_data.append((
                'TEXTCOLOR', (col, row), (col, row), colors.red
            ))

    # Make table
    perolehan_nilai_ilo_table_style = TableStyle(perolehan_nilai_ilo_table_style_data)
    perolehan_nilai_ilo_table = Table(
        perolehan_nilai_ilo_table_data,
        style=perolehan_nilai_ilo_table_style,
        hAlign='LEFT',
        repeatRows=1,
    )

    is_radar_chart = True

    if list_nilai_ilo_peserta.count() <= 2:
        is_radar_chart = False

    if is_radar_chart:
        # If ILO are more than 2, use radar chart
        theta = radar_factory(len(list_nama_ilo), frame='polygon')

        data = [list_nama_ilo]
        labels = ['Satisfactory level', 'Nilai Mahasiswa']

        # Setup data and label
        data.append((
            mk_kurikulum.nama, [
                perolehan_nilai_ilo_chart_data['Satisfactory level'],
                perolehan_nilai_ilo_chart_data['Nilai mahasiswa'],
            ]
        ))
        
        # R axis label
        spoke_labels = data.pop(0)

        # Setup radar chart
        fig, axs = plt.subplots(
            figsize=(3, 3), nrows=1, ncols=1,
            subplot_kw=dict(projection='radar'))
        fig.subplots_adjust(wspace=0.25, hspace=0.20, top=0.85, bottom=0.05)
        line_colors = ['r', 'b']

        # Styling
        for ax, (plot_title, case_data) in zip([axs], data):
            ax.set_rlim(0, 100)
            ax.set_rgrids([0, 20, 40, 60, 80, 100])
            ax.set_title(
                plot_title, weight='bold', size='medium', position=(0.5, 1.1),
                horizontalalignment='center', verticalalignment='center')
            for d, color in zip(case_data, line_colors):
                ax.plot(theta, d, color=color)
            ax.set_varlabels(spoke_labels)
        
        # add legend relative to top-left plot
        axs.legend(
            labels, loc=(0.9, .95),
            labelspacing=0.1, fontsize='small'
        )

        # Chart title
        fig.text(
            0.5, 0.965, 
            'Grafik Perolehan Nilai CPL',
            horizontalalignment='center', 
            color='black', weight='bold',
            size='large'
        )
        
        # Save chart
        perolehan_nilai_ilo_chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
        perolehan_nilai_ilo_chart_path = save_chart(perolehan_nilai_ilo_chart_dir, fig)
    else:
        x = np.arange(len(list_nama_ilo))  # the label locations
        width = 0.25  # the width of the bars
        multiplier = 0

        fig, ax = plt.subplots(layout='constrained')

        # Bar chart
        for nama_filter, list_nilai in perolehan_nilai_ilo_chart_data.items():
            offset = width * multiplier
            ax.bar(
                x + offset, 
                list_nilai, 
                width, 
                label=nama_filter
            )
            multiplier += 1

        # Styling
        ax.set_title('Grafik Perolahan Nilai CPL')
        ax.set_xticks(x + (len(list_nama_ilo)-1) * width/2, list_nama_ilo)
        ax.grid(True, axis='y', which='major')
        ax.grid(True, axis='x', which='minor')
        
        minor_locator = mticker.AutoMinorLocator(2) # set the number of minor intervals per major interval
        ax.xaxis.set_minor_locator(minor_locator)
        ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=2)
        ax.set_ylim(0, 100)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color('#DDDDDD')
        
        # Save chart
        perolehan_nilai_ilo_chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
        perolehan_nilai_ilo_chart_path = save_chart(perolehan_nilai_ilo_chart_dir, fig)

    perolehan_nilai_ilo_chart_image = Image(perolehan_nilai_ilo_chart_path)

    # Build
    pdf_file.build([
        title, detail,
        perolehan_nilai_clo_title,
        perolehan_nilai_clo_table, empty_line,
        perolehan_nilai_ilo_title,
        perolehan_nilai_ilo_table, PageBreak(),
        perolehan_nilai_clo_chart_image, PageBreak(), 
        perolehan_nilai_ilo_chart_image
    ])
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream
