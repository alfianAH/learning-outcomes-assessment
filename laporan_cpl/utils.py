from functools import partial
from io import BytesIO
from copy import copy
import os
import openpyxl
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import QuerySet, Count
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    Paragraph, Spacer, Table, TableStyle,
    Frame, PageTemplate, Image, PageBreak
)
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from learning_outcomes_assessment.utils import export_pdf_header
from learning_outcomes_assessment.chart.utils import (
    radar_factory,
    save_chart
)
from kurikulum.models import Kurikulum
from ilo.models import Ilo
from pi_area.models import PerformanceIndicator
from clo.models import (
    Clo, NilaiCloMataKuliahSemester,
    NilaiCloPeserta,
)
from mata_kuliah_semester.models import (
    MataKuliahSemester,
    PesertaMataKuliah,
    NilaiMataKuliahIloMahasiswa,
)


User = get_user_model()
matplotlib.use('Agg')


def calculate_ilo(list_bobot_mk: list, list_persentase_clo: list,
                  list_bobot_pi_ilo: list, list_nilai_clo: list,
                  nilai_max: int):
    """Calculate ILO

    Args:
        list_bobot_mk (list): list of bobot MK
        list_persentase_clo (list): List of persentase CLO
        list_bobot_pi_ilo (list): List of bobot PI ILO
        list_nilai_clo (list): List of nilai CLO Peserta or MK
        nilai_max (int): Nilai max

    Returns:
        (bool, float): 
            is_success: Is calculation success?
            persentase_capaian_ilo: ILO achievement score
    """

    is_success = False

    list_bobot_mk = np.array(list_bobot_mk)
    list_persentase_clo = np.array(list_persentase_clo)
    list_bobot_pi_ilo = np.array(list_bobot_pi_ilo)
    list_nilai_clo = np.array(list_nilai_clo)

    # Use try to prevent miss shape
    try:
        konversi_clo_to_ilo = (list_bobot_mk * list_persentase_clo * list_bobot_pi_ilo) * nilai_max
    except ValueError:
        message = 'Shape array tidak sama. Bobot MK: {} Persentase CPMK: {}, Bobot PI CPL: {}.'.format(
            list_bobot_mk.shape, list_persentase_clo.shape, list_bobot_pi_ilo.shape)
        if settings.DEBUG: print(message)
        return (is_success, None)
    
    # If mahasiswa doesn't have the ILO, just skip it
    if konversi_clo_to_ilo.shape[0] != 0:
        persentase_konversi = konversi_clo_to_ilo / np.sum(konversi_clo_to_ilo)
    else:
        return (is_success, None)

    # Use try to prevent miss shape
    try:
        persentase_capaian_ilo = persentase_konversi * list_nilai_clo
    except ValueError:
        message = 'Shape array tidak sama. Persentase Konversi: {}, Nilai CPMK: {}.'.format(
            persentase_konversi.shape, list_nilai_clo.shape)
        if settings.DEBUG: print(message)
        return (is_success, None)

    persentase_capaian_ilo = np.sum(persentase_capaian_ilo)
    is_success = True

    return (is_success, persentase_capaian_ilo)


def calculate_ilo_prodi(list_ilo: QuerySet[Ilo], 
                        list_mk_semester: QuerySet[MataKuliahSemester], 
                        max_sks_prodi: int, nilai_max: int):
    """Calculate nilai per ILO from list_ilo and list_mk_semester

    Args:
        list_ilo (QuerySet[Ilo]): Kurikulum's ILO QuerySet
        list_mk_semester (QuerySet[MataKuliahSemester]): Mata Kuliah Semester QuerySet
        max_sks_prodi (int): SKS Max Kelulusan Prodi
        nilai_max (int): Nilai max

    Returns:
        (bool, str, dict): 
            is_success: Is calculation success?
            message: Success or error message
            result: Result of calculation 
            Example:
                result = {
                    'nama_ilo': nilai,
                    'nama_ilo': nilai,
                }
    """

    is_success = False
    message = ''
    result = {}
    calculated_mk_semester = []

    # Loop through ILO
    for ilo in list_ilo:
        # Get all mata kuliah semester
        list_mk_semester_ilo = list_mk_semester.filter(
            clo__piclo__performance_indicator__pi_area__ilo=ilo
        ).distinct()
        
        list_bobot_mk = []
        list_persentase_clo = []
        list_bobot_pi_ilo = []
        list_nilai_clo_mk_semester = []

        # Loop through MK Semester
        for mk_semester in list_mk_semester_ilo:
            # Check status nilai MK Semester
            if not mk_semester.status_nilai:
                peserta_qs = PesertaMataKuliah.objects.filter(kelas_mk_semester__mk_semester=mk_semester).exclude(nilai_akhir=None)
                # Skip mata kuliah that doesn't have nilai akhir yet
                if peserta_qs.count() == 0:
                    continue
                
                message = 'Harap melengkapi nilai dari mata kuliah ({}, {}) terlebih dahulu'.format(
                    mk_semester.mk_kurikulum.nama, mk_semester.semester.semester)
                return (is_success, message, result)

            # Add MK Semester to calculated_mk_semester
            if mk_semester not in calculated_mk_semester:
                calculated_mk_semester.append(mk_semester)

            list_clo_ilo: QuerySet[Clo] = Clo.objects.filter(
                mk_semester=mk_semester,
                piclo__performance_indicator__pi_area__ilo=ilo,
            ).distinct()

            # Get bobot MK
            list_bobot_mk += [mk_semester.mk_kurikulum.sks / max_sks_prodi for _ in range(list_clo_ilo.count())]

            # Get list persentase CLO
            list_persentase_clo += [
                clo.get_total_persentase_komponen()/100 for clo in list_clo_ilo
            ]

            # Get list bobot PI
            list_pi_ilo = PerformanceIndicator.objects.filter(
                pi_area__ilo=ilo
            )
            list_bobot_pi_ilo += [
                clo.get_pi_clo().count() / list_pi_ilo.count() for clo in list_clo_ilo
            ]

            # Get nilai CLO MK
            nilai_clo_mk_semester_qs = NilaiCloMataKuliahSemester.objects.filter(
                mk_semester=mk_semester,
                clo__in=list_clo_ilo
            ).values_list('nilai')

            if nilai_clo_mk_semester_qs.exists():
                list_nilai_clo_mk_semester += [nilai_clo[0] for nilai_clo in nilai_clo_mk_semester_qs]
            else:
                message = 'Mata Kuliah: {}, belum mempunyai Nilai CPMK.'.format(mk_semester.mk_kurikulum.nama)
                return (is_success, message, result)
        
        # Calculate ILO
        is_calculation_success, persentase_capaian_ilo = calculate_ilo(list_bobot_mk, list_persentase_clo, list_bobot_pi_ilo, list_nilai_clo_mk_semester, nilai_max)
        
        # Add to result
        if is_calculation_success:
            result[ilo.nama] = persentase_capaian_ilo
        else:
            result[ilo.nama] = None
            continue
    
    expected_len_mk_semester = list_mk_semester.count()
    # If calculated mk semester is not the same as in query, return
    if len(calculated_mk_semester) != expected_len_mk_semester:
        remaining_len_mk_semester = expected_len_mk_semester - len(calculated_mk_semester)

        message = 'Terdapat {} mata kuliah yang belum dihitung. Total yang sudah dihitung: {}, ekspektasi: {}'.format(remaining_len_mk_semester, len(calculated_mk_semester), expected_len_mk_semester)

        return (is_success, message, result)

    if len(result.keys()) != list_ilo.count():
        message = 'Panjang array List CPL dan hasil tidak sesuai. List CPL: {}, Hasil: {}.'.format(list_ilo.count(), len(result.keys()))
    else:
        is_success = True
        message = 'Berhasil menghitung capaian CPL Program Studi.'

    return (is_success, message, result)


def calculate_ilo_mahasiswa(list_ilo: QuerySet[Ilo], 
                            list_mahasiswa_as_peserta_mk: QuerySet[PesertaMataKuliah],
                            mahasiswa: User, 
                            max_sks_prodi: int, nilai_max: int):
    """Calculate nilai per ILO from list_mahasiswa_as_peserta_mk

    Args:
        list_ilo (QuerySet[Ilo]): Kurikulum's ILO QuerySet
        list_mahasiswa_as_peserta_mk (QuerySet[PesertaMataKuliah]): Peserta MK Queryset
        mahasiswa (User): Current mahasiswa (User)
        max_sks_prodi (int): SKS Max Kelulusan Prodi
        nilai_max (int): Nilai max

    Returns:
        (bool, str, list, list): 
            is_success: Is calculation success?
            message: Success or error message
            result: Result of calculation 
            Example:
                result = [{
                    'nama': 'nama_ilo',
                    'nilai': nilai_ilo,
                    'satisfactory_level: satisfactory_level_ilo
                }]
    """

    is_success = False
    message = ''
    result = []
    calculated_peserta = []
    
    # Loop through ILO
    for ilo in list_ilo:
        list_peserta_mk_ilo = list_mahasiswa_as_peserta_mk.filter(
            kelas_mk_semester__mk_semester__clo__piclo__performance_indicator__pi_area__ilo=ilo
        ).distinct()

        list_bobot_mk = []
        list_persentase_clo = []
        list_bobot_pi_ilo = []
        list_nilai_clo_peserta = []

        # Loop through MK Semester (peserta)
        for peserta_mk in list_peserta_mk_ilo:
            mk_semester = peserta_mk.kelas_mk_semester.mk_semester

            # Check status nilai
            if not peserta_mk.status_nilai:
                message = 'Harap melengkapi nilai dari {} di mata kuliah ({}, {}) terlebih dahulu'.format(
                    mahasiswa.nama, 
                    mk_semester.mk_kurikulum.nama,
                    mk_semester.semester)
                if settings.DEBUG: print(message)
                return (is_success, message, result)

            if peserta_mk not in calculated_peserta:
                calculated_peserta.append(peserta_mk)

            list_clo_ilo = Clo.objects.filter(
                mk_semester=mk_semester,
                piclo__performance_indicator__pi_area__ilo=ilo,
            ).distinct()

            # Get bobot MK
            list_bobot_mk += [mk_semester.mk_kurikulum.sks / max_sks_prodi for _ in range(list_clo_ilo.count())]

            # Get list persentase CLO
            list_persentase_clo += [
                clo.get_total_persentase_komponen()/100 for clo in list_clo_ilo
            ]

            # Get list bobot PI
            list_pi_ilo = PerformanceIndicator.objects.filter(
                pi_area__ilo=ilo
            )
            list_bobot_pi_ilo += [
                clo.get_pi_clo().count() / list_pi_ilo.count() for clo in list_clo_ilo
            ]

            # Get nilai CLO peserta
            nilai_clo_peserta_qs = NilaiCloPeserta.objects.filter(
                peserta=peserta_mk,
                clo__in=list_clo_ilo
            ).values_list('nilai')

            if nilai_clo_peserta_qs.exists():
                list_nilai_clo_peserta += [nilai_clo[0] for nilai_clo in nilai_clo_peserta_qs]
            else:
                message = 'Peserta: {} di Mata Kuliah: {}, {}, belum mempunyai Nilai CPMK.'.format(mahasiswa.nama, mk_semester.mk_kurikulum.nama, mk_semester.semester.semester)
                if settings.DEBUG: print(message)
                return (is_success, message, result)

        # Calculate ILO
        is_success, persentase_capaian_ilo = calculate_ilo(list_bobot_mk, list_persentase_clo, list_bobot_pi_ilo, list_nilai_clo_peserta, nilai_max)

        # Add to result
        if is_success:
            result.append({
                'nama': ilo.nama,
                'nilai': persentase_capaian_ilo,
                'satisfactory_level': ilo.satisfactory_level,
            })
        else:
            result.append({
                'nama': ilo.nama,
                'nilai': None,
                'satisfactory_level': ilo.satisfactory_level,
            })
            continue
    
    expected_len_peserta_mk = list_mahasiswa_as_peserta_mk.count()
    # If calculated mk semester is not the same as in query, return
    if len(calculated_peserta) != expected_len_peserta_mk:
        remaining_len_peserta = expected_len_peserta_mk - len(calculated_peserta)

        message = 'Terdapat {} peserta yang belum dihitung. Total yang sudah dihitung: {}, ekspektasi: {}'.format(remaining_len_peserta, len(calculated_peserta), expected_len_peserta_mk)

        return (is_success, message, result)
    
    is_success = True
    message = 'Berhasil menghitung capaian CPL Mahasiswa.'
    return (is_success, message, result)


def perolehan_nilai_ilo_graph(list_ilo, is_multiple_result):
    perolehan_nilai_ilo_graph_dict = {
        'labels': [ilo.nama for ilo in list_ilo],
        'datasets': [
            {
                'label': 'Satisfactory Level',
                'data': [ilo.satisfactory_level for ilo in list_ilo],
                'fill': False,
            }
        ]
    }

    if is_multiple_result:
        # Update to bar chart
        perolehan_nilai_ilo_graph_dict.update({
            'chart_type': 'bar'
        })
        # Update satisfactory level to line
        perolehan_nilai_ilo_graph_dict['datasets'][0].update({
            'type': 'line'
        })
    else:
        perolehan_nilai_ilo_graph_dict.update({
            'chart_type': 'radar'
        })

    return perolehan_nilai_ilo_graph_dict


def perolehan_nilai_ilo_prodi_graph(calculation_result: dict, json_dict: dict):
    for key, value in calculation_result.items():
        json_dict['datasets'].append({
            'label': key,
            'data': [nilai_ilo for nama_ilo, nilai_ilo in value.items()],
            'fill': False,
        })


def perolehan_nilai_ilo_mahasiswa_graph(calculation_result: dict, json_dict: dict):
    for _, data_mhs in calculation_result.items():
        for filter in data_mhs['result']:
            filter_name = filter['filter']
            filter_results = filter['result']

            json_dict['datasets'].append({
                'label': filter_name,
                'data': [filter_result['nilai'] for filter_result in filter_results],
                'fill': False,
            })

        break


def get_ilo_and_sks_from_kurikulum(kurikulum_obj: Kurikulum):
    """Get ILO and max SKS prodi

    Args:
        kurikulum_id (str): Kurikulum ID

    Returns:
        (QuerySet[ILO], int): QuerySet of ILO and max SKS Prodi
    """

    list_ilo: QuerySet[Ilo] = Ilo.objects.filter(
        pi_area__assessment_area__kurikulum=kurikulum_obj
    )

    max_sks_prodi = kurikulum_obj.prodi_jenjang.total_sks_lulus

    return (list_ilo, max_sks_prodi)


def generate_laporan_cpl_chart(
        chart_data: dict,
        list_ilo: QuerySet[Ilo], list_filter: list,
        chart_title: str,
):
    """Generate chart for laporan CPL

    Args:
        chart_data (dict): Chart data
        list_ilo (QuerySet[Ilo]): List ILO
        list_filter (list): List filter laporan CPL
        chart_title (str): Chart title

    Returns:
        str: Chart saved location path
    """
    
    is_multiple_result = len(list_filter) > 1
    list_nama_ilo = []
    satisfactory_level_data = []

    for ilo in list_ilo:
        list_nama_ilo.append(ilo.nama)
        satisfactory_level_data.append(ilo.satisfactory_level)
    
    # If multiple result, bar and line chart
    # Else, radar chart
    if is_multiple_result:
        x = np.arange(len(list_nama_ilo))  # the label locations
        width = 0.25  # the width of the bars
        multiplier = 0

        fig, ax = plt.subplots(layout='constrained')

        # Bar chart
        for nama_filter, list_nilai_ilo in chart_data.items():
            offset = width * multiplier
            ax.bar(
                x + offset, 
                list_nilai_ilo, 
                width, 
                label=nama_filter
            )
            multiplier += 1

        # Satisfactory level line chart
        ax.plot(
            x + (len(list_filter)-1) * width/2, 
            satisfactory_level_data, 
            marker='o', 
            color='dimgrey', 
            label='Satisfactory level'
        )

        # Styling
        ax.set_title(chart_title)
        ax.set_xticks(x + (len(list_filter)-1) * width/2, list_nama_ilo)
        ax.grid(True, axis='y', which='major')
        ax.grid(True, axis='x', which='minor')
        
        minor_locator = mticker.AutoMinorLocator(2) # set the number of minor intervals per major interval
        ax.xaxis.set_minor_locator(minor_locator)
        ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=2)
        ax.set_ylim(0, 100)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color('#DDDDDD')
        
        # Save chart
        chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
        chart_path = save_chart(chart_dir, fig)
    else:
        if list_ilo.count() > 2:
            # If ILO are more than 2, use radar chart
            theta = radar_factory(list_ilo.count(), frame='polygon')

            data = [list_nama_ilo]
            labels = ['Satisfactory level']

            # Setup data and label
            for nama_filter, list_nilai_ilo in chart_data.items():
                labels.append(nama_filter)
                data.append((
                    nama_filter, [
                        satisfactory_level_data,
                        list_nilai_ilo,
                    ]
                ),)

            # R axis label
            spoke_labels = data.pop(0)

            # Setup radar chart
            fig, axs = plt.subplots(
                figsize=(3, 3), nrows=1, ncols=1,
                subplot_kw=dict(projection='radar'))
            fig.subplots_adjust(wspace=0.25, hspace=0.20, top=0.85, bottom=0.05)
            line_colors = ['r', 'b']

            # Styling
            for ax, (plot_title, case_data) in zip([axs], data):
                ax.set_rlim(0, 100)
                ax.set_rgrids([0, 20, 40, 60, 80, 100])
                ax.set_title(
                    plot_title, weight='bold', size='medium', position=(0.5, 1.1),
                    horizontalalignment='center', verticalalignment='center')
                for d, color in zip(case_data, line_colors):
                    ax.plot(theta, d, color=color, marker='o')
                ax.set_varlabels(spoke_labels)
            
            # add legend relative to top-left plot
            axs.legend(
                labels, loc=(0.9, .95),
                labelspacing=0.1, fontsize='small'
            )

            # Chart title
            fig.text(
                0.5, 0.965, 
                chart_title,
                horizontalalignment='center', 
                color='black', weight='bold',
                size='large'
            )
            
            # Save chart
            chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
            chart_path = save_chart(chart_dir, fig)
        else:
            # If ILO are less or equal to 2, use bar chart
            x = np.arange(len(list_nama_ilo))  # the label locations
            width = 0.25  # the width of the bars
            multiplier = 0

            fig, ax = plt.subplots(layout='constrained')
            chart_data['Satisfactory level'] = satisfactory_level_data
            # Bar chart
            for nama_filter, list_nilai_ilo in chart_data.items():
                offset = width * multiplier
                ax.bar(
                    x + offset, 
                    list_nilai_ilo, 
                    width, 
                    label=nama_filter
                )
                multiplier += 1

            # Styling
            ax.set_title(chart_title)
            ax.set_xticks(x + width/2, list_nama_ilo)
            ax.grid(True, axis='y', which='major')
            ax.grid(True, axis='x', which='minor')
            
            minor_locator = mticker.AutoMinorLocator(2) # set the number of minor intervals per major interval
            ax.xaxis.set_minor_locator(minor_locator)
            ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=2)
            ax.set_ylim(0, 100)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_visible(False)
            ax.spines['bottom'].set_color('#DDDDDD')
            
            # Save chart
            chart_dir =  os.path.join(settings.STATIC_ROOT, 'laporan_cpl', 'charts')
            chart_path = save_chart(chart_dir, fig)

    return chart_path


def generate_keterangan_ilo(list_ilo: QuerySet[Ilo]):
    styles = getSampleStyleSheet()

    # (COL, ROW)
    table_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]

    # Keterangan ILO
    ilo_detail_title = Paragraph(
        'Detail Capaian Pembelajaran Lulusan',
        style=styles['h2']
    )
    ilo_detail_table_data = [
        ['No.', 'CPL', 'Deskripsi']
    ]

    for i, ilo in enumerate(list_ilo):
        ilo_detail_table_data.append((
            '{}.'.format(i+1), 
            ilo.nama, 
            Paragraph('{}'.format(ilo.deskripsi))
        ))

    ilo_detail_table = Table(
        ilo_detail_table_data,
        style=table_style_data,
        hAlign='LEFT',
        colWidths=[1*cm, 3*cm, None],
        repeatRows=1,
    )

    return ilo_detail_title, ilo_detail_table


def generate_laporan_cpl_prodi_pdf(
        list_ilo: QuerySet[Ilo],
        list_filter,
        calculation_result: dict,
        prodi_name: str, fakultas_name: str,
):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    page_size = landscape(letter)
    pdf_file = SimpleDocTemplate(file_stream, pagesize=page_size)
    styles = getSampleStyleSheet()
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        'Laporan Capaian Pembelajaran Lulusan Prodi',
        style=styles['h1']
    )

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas_name, prodi_name), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])

    # Tabel detail spider chart
    table_spider_chart_title = Paragraph(
        'Tabel Pencapaian Capaian Pembelajaran Lulusan',
        style=styles['h2']
    )

    # (COL, ROW)
    table_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]
    
    # Table spider chart
    len_table_spider_chart = len(list_filter) // 3
    if len(list_filter) % 3 > 0: len_table_spider_chart += 1
    
    list_table_spider_chart = []
    chart_data = {}
    
    for i in range(len_table_spider_chart):
        table_data = [['CPL', 'Satisfactory Level',]]
        current_table_style_data = copy(table_style_data)
        
        # Calculate index subscriptable
        filter_index_start = i*3

        if len_table_spider_chart > 1:
            # If table is more than 1, ...
            if i+1 == len_table_spider_chart:
                filter_index_end = filter_index_start + len(list_filter) % 3
            else:
                filter_index_end = filter_index_start + 3
        else:
            # If table is only 1 row, ...
            filter_index_end = len(list_filter)
        
        # Make header for filter
        for filter in list_filter[filter_index_start:filter_index_end]:
            table_data[0].append(filter[1])

        # Table spider chart data
        for j, ilo in enumerate(list_ilo):
            data = [ilo.nama, ilo.satisfactory_level]

            calcultation_result_sub = list(calculation_result.items())[filter_index_start:filter_index_end]
            for k, (nama_filter, ilo_dict) in enumerate(calcultation_result_sub):
                # Set chart data
                if nama_filter not in chart_data.keys():
                    chart_data[nama_filter] = []
                
                for nama_ilo, nilai_ilo in ilo_dict.items():
                    # Skip if not current ILO
                    if nama_ilo != ilo.nama: continue

                    # Add nilai to row
                    if nilai_ilo is None: 
                        data.append('-')
                        # Add chart data
                        chart_data[nama_filter].append(np.NaN)
                    else: 
                        nilai_ilo_formatted = float("{:.2f}".format(nilai_ilo))
                        data.append(nilai_ilo_formatted)
                        # Add chart data
                        chart_data[nama_filter].append(nilai_ilo_formatted)

                        # Give red color
                        if nilai_ilo_formatted < ilo.satisfactory_level:
                            col = 2 + k
                            row = 1 + j
                            current_table_style_data.append((
                                'TEXTCOLOR', (col, row), (col, row), colors.red
                            ))
                    
                    # Go to next filter
                    break

            # Append data
            table_data.append(data)

        current_table_spider_chart_style = TableStyle(current_table_style_data)
        spider_chart_table = Table(
            table_data,
            style=current_table_spider_chart_style,
            hAlign='LEFT',
            repeatRows=1,
        )

        list_table_spider_chart.append(spider_chart_table)
    
    # Chart ILO
    chart_path = generate_laporan_cpl_chart(
        chart_data,
        list_ilo, list_filter, 
        'Laporan Capaian Pembelajaran Lulusan Program Studi',
    )
    chart_image = Image(chart_path)

    # Keterangan ILO
    ilo_detail_title, ilo_detail_table = generate_keterangan_ilo(list_ilo)

    # Build
    pdf_file_elements = [
        title,
        table_spider_chart_title,
    ]
    for spider_chart_table in list_table_spider_chart:
        pdf_file_elements += [spider_chart_table, empty_line]
    
    pdf_file_elements += [
        PageBreak(), chart_image, PageBreak(),
        ilo_detail_title,
        ilo_detail_table,
    ]
    pdf_file.build(pdf_file_elements)
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream


def generate_all_bobot_perhitungan_cpl_prodi(
        current_worksheet,
        list_mk_semester: QuerySet[MataKuliahSemester],
        ilo: Ilo,
        max_sks_prodi: int,
):
    """Generate bobot perhitungan CPL mahasiswa for raw report

    Args:
        current_worksheet (Worksheet): worksheet
        list_mk_semester (QuerySet[MataKuliahSemester]): list mata kuliah semester
        ilo (Ilo): ILO
        max_sks_prodi (int): Max SKS kelulusan Prodi
    """

    start_row = 2

    # Loop through MK Semester
    for mk_semester in list_mk_semester:
        # Check status nilai MK Semester
        if not mk_semester.status_nilai:
            peserta_qs = PesertaMataKuliah.objects.filter(kelas_mk_semester__mk_semester=mk_semester).exclude(nilai_akhir=None)
            # Skip mata kuliah that doesn't have nilai akhir yet
            if peserta_qs.count() == 0: continue
            
            continue

        list_clo_ilo: QuerySet[Clo] = Clo.objects.filter(
            mk_semester=mk_semester,
            piclo__performance_indicator__pi_area__ilo=ilo,
        ).distinct()

        # Get list bobot PI
        list_pi_ilo = PerformanceIndicator.objects.filter(
            pi_area__ilo=ilo
        )

        for clo in list_clo_ilo:
            persentase_clo = clo.get_total_persentase_komponen()/100
            bobot_pi_ilo = clo.get_pi_clo().count() / list_pi_ilo.count()

            # Get nilai CLO MK
            nilai_clo_mk_semester_qs = NilaiCloMataKuliahSemester.objects.filter(
                mk_semester=mk_semester,
                clo=clo
            ).values_list('nilai')

            # Excel
            current_worksheet['A{}'.format(start_row)] = mk_semester.mk_kurikulum.kode
            current_worksheet['B{}'.format(start_row)] = mk_semester.mk_kurikulum.nama
            current_worksheet['C{}'.format(start_row)] = clo.nama
            current_worksheet['D{}'.format(start_row)] = '={}/{}'.format(mk_semester.mk_kurikulum.sks, max_sks_prodi)
            current_worksheet['E{}'.format(start_row)] = persentase_clo
            current_worksheet['F{}'.format(start_row)] = bobot_pi_ilo
            current_worksheet['G{}'.format(start_row)] = 100

            if nilai_clo_mk_semester_qs.exists():
                nilai_clo_mk_semester = nilai_clo_mk_semester_qs.first()[0]
                current_worksheet['H{}'.format(start_row)] = nilai_clo_mk_semester
            else:
                current_worksheet['H{}'.format(start_row)] = 0
                continue
            
            start_row += 1


def generate_all_bobot_perhitungan_cpl_mahasiswa(
        current_worksheet,
        list_peserta_mk_semester: QuerySet[PesertaMataKuliah],
        ilo: Ilo,
        max_sks_prodi: int,
):
    """Generate bobot perhitungan CPL mahasiswa for raw report

    Args:
        current_worksheet (Sheet): Worksheet
        list_peserta_mk_semester (QuerySet[PesertaMataKuliah]): Peserta Mata Kuliah
        ilo (Ilo): ILO
        max_sks_prodi (int): Max SKS kelulusan Prodi
    """
    start_row = 2

    # Loop through MK Semester
    for peserta_mk_semester in list_peserta_mk_semester:
        mk_semester = peserta_mk_semester.kelas_mk_semester.mk_semester

        # Check status nilai MK Semester
        if not peserta_mk_semester.status_nilai:
            continue

        list_clo_ilo: QuerySet[Clo] = Clo.objects.filter(
            mk_semester=mk_semester,
            piclo__performance_indicator__pi_area__ilo=ilo,
        ).distinct()

        # Get list bobot PI
        list_pi_ilo = PerformanceIndicator.objects.filter(
            pi_area__ilo=ilo
        )

        for clo in list_clo_ilo:
            persentase_clo = clo.get_total_persentase_komponen()/100
            bobot_pi_ilo = clo.get_pi_clo().count() / list_pi_ilo.count()

            # Get nilai CLO peserta
            nilai_clo_peserta_qs = NilaiCloPeserta.objects.filter(
                peserta=peserta_mk_semester,
                clo=clo
            ).values_list('nilai')

            # Excel
            current_worksheet['A{}'.format(start_row)] = mk_semester.mk_kurikulum.kode
            current_worksheet['B{}'.format(start_row)] = mk_semester.mk_kurikulum.nama
            current_worksheet['C{}'.format(start_row)] = clo.nama
            current_worksheet['D{}'.format(start_row)] = '={}/{}'.format(mk_semester.mk_kurikulum.sks, max_sks_prodi)
            current_worksheet['E{}'.format(start_row)] = persentase_clo
            current_worksheet['F{}'.format(start_row)] = bobot_pi_ilo
            current_worksheet['G{}'.format(start_row)] = 100

            if nilai_clo_peserta_qs.exists():
                nilai_clo_peserta_mk_semester = nilai_clo_peserta_qs.first()[0]
                current_worksheet['H{}'.format(start_row)] = nilai_clo_peserta_mk_semester
            else:
                current_worksheet['H{}'.format(start_row)] = 0
                continue
            
            start_row += 1


def generate_laporan_cpl_prodi_raw(
        list_ilo: QuerySet[Ilo], max_sks_prodi: int, 
        list_filter, 
        is_kurikulum: bool = False, is_semester_included: bool = False
):
    """Generate laporan CPL Prodi Raw

    Args:
        list_ilo (QuerySet[Ilo]): List ILO
        max_sks_prodi (int): Max SKS kelulusan prodi
        list_filter (list): Filter kurikulum/tahun ajaran/semester
        is_kurikulum (bool, optional): is filter kurikulum only?. Defaults to False.
        is_semester_included (bool, optional): does filter include semester?. Defaults to False.

    Returns:
        FileStream: excel file
    """

    workbook = openpyxl.Workbook()

    # Remove default sheet
    workbook.remove_sheet(workbook['Sheet'])

    # Create sheets
    for ilo in list_ilo:
        workbook.create_sheet(ilo.nama)
        current_worksheet = workbook[ilo.nama]

        current_worksheet['A1'] = 'Kode MK'
        current_worksheet['B1'] = 'Nama MK'
        current_worksheet['C1'] = 'CPMK'
        current_worksheet['D1'] = 'Bobot MK'
        current_worksheet['E1'] = 'Bobot CPMK'
        current_worksheet['F1'] = 'Bobot PI CPL'
        current_worksheet['G1'] = 'Nilai maksimum'
        current_worksheet['H1'] = 'Nilai CPMK'

        if is_kurikulum:
            kurikulum_obj: Kurikulum = list_filter[0][0]
            # Get all mata kuliah semester
            list_mk_semester = MataKuliahSemester.objects.annotate(
                num_peserta=Count('kelasmatakuliahsemester__pesertamatakuliah')
            ).filter(
                num_peserta__gt=0,
                mk_kurikulum__kurikulum=kurikulum_obj,
                clo__piclo__performance_indicator__pi_area__ilo=ilo,
                kelasmatakuliahsemester__pesertamatakuliah__nilai_akhir__isnull=False
            ).distinct()

            generate_all_bobot_perhitungan_cpl_prodi(current_worksheet, list_mk_semester, ilo, max_sks_prodi)
        else:
            if is_semester_included:
                list_semester_obj = []
                for semester_prodi_obj, semester_nama in list_filter:
                    list_semester_obj.append(semester_prodi_obj)
                
                # Get all mata kuliah semester
                list_mk_semester = MataKuliahSemester.objects.annotate(
                    num_peserta=Count('kelasmatakuliahsemester__pesertamatakuliah')
                ).filter(
                    num_peserta__gt=0,
                    semester__in=list_semester_obj,
                    clo__piclo__performance_indicator__pi_area__ilo=ilo,
                    kelasmatakuliahsemester__pesertamatakuliah__nilai_akhir__isnull=False
                ).distinct()

                generate_all_bobot_perhitungan_cpl_prodi(current_worksheet, list_mk_semester, ilo, max_sks_prodi)
            else:
                list_tahun_ajaran_obj = []
                for tahun_ajaran_prodi_obj, tahun_ajaran_nama in list_filter:
                    list_tahun_ajaran_obj.append(tahun_ajaran_prodi_obj)
                
                # Get all mata kuliah semester
                list_mk_semester = MataKuliahSemester.objects.annotate(
                    num_peserta=Count('kelasmatakuliahsemester__pesertamatakuliah')
                ).filter(
                    num_peserta__gt=0,
                    semester__tahun_ajaran_prodi__in=list_tahun_ajaran_obj,
                    kelasmatakuliahsemester__pesertamatakuliah__nilai_akhir__isnull=False
                ).distinct()

                generate_all_bobot_perhitungan_cpl_prodi(current_worksheet, list_mk_semester, ilo, max_sks_prodi)

    # Save the workbook to a BytesIO object
    file_stream = BytesIO()
    workbook.save(file_stream)

    file_stream.seek(0)

    return file_stream


def generate_laporan_cpl_mahasiswa_pdf(
        list_ilo: QuerySet[Ilo],
        list_filter,
        calculation_result: dict,
        prodi_name: str, fakultas_name: str,
):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    page_size = landscape(letter)
    pdf_file = SimpleDocTemplate(file_stream, pagesize=page_size)
    styles = getSampleStyleSheet()
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        'Laporan Capaian Pembelajaran Lulusan Mahasiswa',
        style=styles['h1']
    )

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas_name, prodi_name), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])

    # Tabel detail ILO
    table_detail_ilo_title = Paragraph(
        'Tabel Pencapaian Capaian Pembelajaran Lulusan',
        style=styles['h2']
    )

    # (COL, ROW)
    table_detail_ilo_style_data = [
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 1), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 1), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 1), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
        ('SPAN', (0, 0), (0, 1)),                           # Nomor merge
        ('SPAN', (1, 0), (1, 1)),                           # NIM merge
        ('SPAN', (2, 0), (2, 1)),                           # Nama merge
        ('VALIGN', (0, 0), (2, 0), 'MIDDLE'),               # Header vertical align
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('ALIGN', (3, 2), (-1, -1), 'RIGHT'),               # Nilai align
    ]

    list_table_detail_ilo = []
    
    for filter in list_filter:
        table_data = [['No.', 'NIM', 'Nama', ]]
        # Add filter name to header
        table_data[0].append(
            Paragraph(
                filter[1],
                style=ParagraphStyle(
                    name='table_filter_header',
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold',
                    fontSize=11
                )
            )
        )

        current_table_detail_ilo_style_data = copy(table_detail_ilo_style_data)

        current_table_detail_ilo_style_data += [
            ('SPAN', (3, 0), (-1, 0)),           # Merge filter name cells
        ]
        
        # Space for Nomor NIM Nama
        ilo_header = ['', '', '']
        ilo_header += [ilo.nama for ilo in list_ilo]
        table_data.append(ilo_header)

        # List mahasiswa
        for i, (nim, peserta_mk) in enumerate(calculation_result.items()):
            list_peserta_ilo_result: list[dict] = peserta_mk['result']
            mahasiswa_row = [
                '{}.'.format(i+1),
                nim,
                Paragraph(peserta_mk['nama'])
            ]
            for peserta_ilo_result in list_peserta_ilo_result:
                # Check by nama filter
                if filter[1] != peserta_ilo_result['filter']: continue
                
                # Add result to row
                list_ilo_result: list[dict] = peserta_ilo_result['result']
                for j, ilo_result in enumerate(list_ilo_result):
                    nilai_ilo_mahasiswa = ilo_result['nilai']

                    # Append nilai ILO mahasiswa
                    if nilai_ilo_mahasiswa is None:
                        mahasiswa_row.append('-')
                    else:
                        nilai_ilo_mahasiswa = float('{:.2f}'.format(nilai_ilo_mahasiswa))
                        mahasiswa_row.append(nilai_ilo_mahasiswa)

                        # Give red color
                        if nilai_ilo_mahasiswa < ilo_result['satisfactory_level']:
                            col = 3 + j
                            row = 2 + i
                            current_table_detail_ilo_style_data.append((
                                'TEXTCOLOR', (col, row), (col, row), colors.red
                            ))
                    
                # Break after filter is found
                break

            # Add row to table data
            table_data.append(mahasiswa_row)
        
        current_table_detail_ilo_style = TableStyle(current_table_detail_ilo_style_data)
        # Make table
        detail_ilo_table = Table(
            table_data, 
            style=current_table_detail_ilo_style,
            hAlign='LEFT',
            colWidths=[1*cm, 3*cm, 6.5*cm, None],
            repeatRows=(0, 1),
        )

        list_table_detail_ilo.append(detail_ilo_table)

    # Keterangan ILO
    ilo_detail_title, ilo_detail_table = generate_keterangan_ilo(list_ilo)

    # Build
    pdf_file_elements = [
        title,
        table_detail_ilo_title,
    ]
    for detail_ilo_table in list_table_detail_ilo:
        pdf_file_elements += [detail_ilo_table, empty_line]

    pdf_file_elements += [
        PageBreak(),
        ilo_detail_title, ilo_detail_table,
    ]
    
    pdf_file.build(pdf_file_elements)
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream


def generate_laporan_cpl_per_mahasiswa_pdf(
        list_ilo: QuerySet[Ilo],
        list_peserta_mk_semester: QuerySet[PesertaMataKuliah],
        list_filter,
        calculation_result: dict, user,
        prodi_name: str, fakultas_name: str
):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    page_size = landscape(letter)
    pdf_file = SimpleDocTemplate(file_stream, pagesize=page_size)
    styles = getSampleStyleSheet()
    empty_line = Spacer(1, 20)

    # Title
    title = Paragraph(
        'Laporan Capaian Pembelajaran Lulusan Mahasiswa',
        style=styles['h1']
    )

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas_name, prodi_name), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])
    
    user_data = [
        ['Nama', ':', user.nama],
        ['NIM', ':', user.username],
        ['Program Studi', ':', user.prodi],
    ]
    user_table_style = TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
    ])
    user_table = Table(
        data=user_data,
        style=user_table_style,
        hAlign='LEFT'
    )

    # Tabel detail MK dan ILO
    table_detail_mk_ilo_title = Paragraph(
        'Tabel Pencapaian Capaian Pembelajaran Lulusan per Mata Kuliah',
        style=styles['h2']
    )

    # (COL, ROW)
    table_detail_mk_ilo_style_data = [
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 1), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 1), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 1), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
        ('SPAN', (0, 0), (0, 1)),                           # Nomor merge
        ('SPAN', (1, 0), (1, 1)),                           # Kode merge
        ('SPAN', (2, 0), (2, 1)),                           # Nama MK merge
        ('SPAN', (3, 0), (3, 1)),                           # SKS MK merge
        ('SPAN', (4, 0), (4, 1)),                           # Nilai akhir MK merge
        ('SPAN', (5, 0), (-1, 0)),                          # CPL merge
        ('VALIGN', (0, 0), (4, 0), 'MIDDLE'),               # Header vertical align
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('ALIGN', (3, 2), (-1, -1), 'RIGHT'),               # Number align
    ]
    
    # Tabel detail MK dan ILO header
    table_detail_mk_ilo_data = [
        ['No.', 'Kode', 'Nama', 'SKS', 'Nilai akhir', 'CPL'],
        ['', '', '', '', '']
    ]

    # ILO Header
    table_detail_mk_ilo_data[1] += [ilo.nama for ilo in list_ilo]

    # List mk semester
    for i, peserta_mk_semester in enumerate(list_peserta_mk_semester):
        mk_kurikulum = peserta_mk_semester.kelas_mk_semester.mk_semester.mk_kurikulum
        # Detail MK
        mk_semester_row = [
            '{}.'.format(i+1),
            mk_kurikulum.kode,
            Paragraph(mk_kurikulum.nama),
            mk_kurikulum.sks,
        ]

        # Add Nilai akhir
        if peserta_mk_semester.nilai_akhir is None:
            mk_semester_row.append('-')
        else:
            mk_semester_row.append(float('{:.2f}'.format(peserta_mk_semester.nilai_akhir)))

        # Add nilai ILO
        for j, ilo in enumerate(list_ilo):
            nilai_mk_ilo_peserta = NilaiMataKuliahIloMahasiswa.objects.filter(
                peserta=peserta_mk_semester,
                ilo=ilo,
            )

            if nilai_mk_ilo_peserta.exists():
                nilai_ilo_peserta = float('{:.2f}'.format(nilai_mk_ilo_peserta.first().nilai_ilo))
                mk_semester_row.append(nilai_ilo_peserta)

                # Give red color
                if nilai_ilo_peserta < ilo.satisfactory_level:
                    col = 5 + j
                    row = 2 + i
                    table_detail_mk_ilo_style_data.append((
                        'TEXTCOLOR', (col, row), (col, row), colors.red
                    ))
            else:
                mk_semester_row.append('-')
        
        table_detail_mk_ilo_data.append(mk_semester_row)

    # Make table
    table_detail_mk_ilo_style = TableStyle(table_detail_mk_ilo_style_data)
    detail_mk_ilo_table = Table(
        table_detail_mk_ilo_data,
        style=table_detail_mk_ilo_style,
        hAlign='LEFT',
        colWidths=[1*cm, None],
        repeatRows=(0, 1),
    )

    # Table chart header
    table_chart_title = Paragraph(
        'Tabel Pencapaian Capaian Pembelajaran Lulusan',
        style=styles['h2']
    )

    # Table chart
    table_chart_style_data = [
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 0), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
    ]

    # Table spider chart
    len_table_chart = len(list_filter) // 3
    if len(list_filter) % 3 > 0: len_table_chart += 1
    
    list_table_chart = []
    chart_data = {}
    
    for i in range(len_table_chart):
        table_data = [['CPL', 'Satisfactory Level',]]
        current_table_chart_style_data = copy(table_chart_style_data)
        
        # Calculate index subscriptable
        filter_index_start = i*3

        if len_table_chart > 1:
            # If table is more than 1, ...
            if i+1 == len_table_chart:
                filter_index_end = filter_index_start + len(list_filter) % 3
            else:
                filter_index_end = filter_index_start + 3
        else:
            # If table is only 1 row, ...
            filter_index_end = len(list_filter)
        
        # Make header for filter
        for filter in list_filter[filter_index_start:filter_index_end]:
            table_data[0].append(filter[1])

        # Table spider chart data
        for j, ilo in enumerate(list_ilo):
            ilo_row = [ilo.nama, ilo.satisfactory_level]
            
            for _, peserta_mk in calculation_result.items():
                list_peserta_ilo_result: list[dict] = peserta_mk['result']
                
                for k, peserta_ilo_result in enumerate(list_peserta_ilo_result):
                    nama_filter = peserta_ilo_result['filter']
                    if nama_filter not in chart_data.keys():
                        chart_data[nama_filter] = []
                    
                    # Add result to row
                    list_ilo_result: list[dict] = peserta_ilo_result['result']

                    if len(list_ilo_result) == 0:
                        chart_data[nama_filter] = [np.NaN for _ in list_ilo]
                    elif len(list_ilo_result) != len(list_ilo):
                        if settings.DEBUG:
                            print('ILO di result ({}) tidak sama dengan List ILO yang ada ({}).'.format(len(list_ilo_result), len(list_ilo)))

                    for ilo_result in list_ilo_result:
                        if ilo_result['nama'] != ilo.nama: 
                            continue
                        
                        nilai_ilo = ilo_result['nilai']

                        if nilai_ilo is None:
                            ilo_row.append('-')
                            # Add chart data
                            chart_data[nama_filter].append(np.NaN)
                        else:
                            nilai_ilo_formatted = float('{:.2f}'.format(nilai_ilo))

                            ilo_row.append(nilai_ilo_formatted)
                            # Add chart data
                            chart_data[nama_filter].append(nilai_ilo_formatted)

                            # Give red color
                            if nilai_ilo_formatted < ilo_result['satisfactory_level']:
                                col = 2 + k
                                row = 1 + j
                                current_table_chart_style_data.append((
                                    'TEXTCOLOR', (col, row), (col, row), colors.red
                                ))

            table_data.append(ilo_row)

        current_table_chart_style = TableStyle(current_table_chart_style_data)
        
        chart_table = Table(
            table_data,
            style=current_table_chart_style,
            hAlign='LEFT',
            repeatRows=1,
        )

        list_table_chart.append(chart_table)

    # Chart ILO
    chart_path = generate_laporan_cpl_chart(
        chart_data,
        list_ilo, list_filter, 
        'Laporan Capaian Pembelajaran Lulusan Mahasiswa',
    )
    chart_image = Image(chart_path)

    # Keterangan ILO
    ilo_detail_title, ilo_detail_table = generate_keterangan_ilo(list_ilo)

    # Build
    pdf_file_elements = [
        title, user_table,
        table_chart_title,
    ]

    for chart_table in list_table_chart:
        pdf_file_elements += [chart_table, empty_line]
    
    pdf_file_elements += [
        PageBreak(), chart_image, PageBreak(),
        table_detail_mk_ilo_title,
        detail_mk_ilo_table, empty_line,
        PageBreak(),
        ilo_detail_title, ilo_detail_table
    ]
    
    pdf_file.build(pdf_file_elements)
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream


def generate_laporan_cpl_per_mahasiswa_raw(
        list_ilo: QuerySet[Ilo], max_sks_prodi: int, 
        list_peserta_mk: QuerySet[PesertaMataKuliah],
        list_filter, 
        is_kurikulum: bool = False, is_semester_included: bool = False
):
    """Generate laporan CPL mahasiswa Raw

    Args:
        list_ilo (QuerySet[Ilo]): List ILO
        max_sks_prodi (int): Max SKS kelulusan prodi
        list_peserta_mk (QuerySet[PesertaMataKuliah]): list peserta mk
        list_filter (list): Filter kurikulum/tahun ajaran/semester
        is_kurikulum (bool, optional): is filter kurikulum only?. Defaults to False.
        is_semester_included (bool, optional): does filter include semester?. Defaults to False.

    Returns:
        FileStream: excel file
    """

    workbook = openpyxl.Workbook()

    # Remove default sheet
    workbook.remove_sheet(workbook['Sheet'])

    # Create sheets
    for ilo in list_ilo:
        workbook.create_sheet(ilo.nama)
        current_worksheet = workbook[ilo.nama]

        current_worksheet['A1'] = 'Kode MK'
        current_worksheet['B1'] = 'Nama MK'
        current_worksheet['C1'] = 'CPMK'
        current_worksheet['D1'] = 'Bobot MK'
        current_worksheet['E1'] = 'Bobot CPMK'
        current_worksheet['F1'] = 'Bobot PI CPL'
        current_worksheet['G1'] = 'Nilai maksimum'
        current_worksheet['H1'] = 'Nilai CPMK'

        if is_kurikulum:
            kurikulum_obj: Kurikulum = list_filter[0][0]
            # Get mahasiswa's mata kuliah participation
            list_mahasiswa_as_peserta_mk = list_peserta_mk.filter(
                kelas_mk_semester__mk_semester__mk_kurikulum__kurikulum=kurikulum_obj,
                kelas_mk_semester__mk_semester__clo__piclo__performance_indicator__pi_area__ilo=ilo,
                nilai_akhir__isnull=False,
            ).distinct()

            generate_all_bobot_perhitungan_cpl_mahasiswa(current_worksheet, list_mahasiswa_as_peserta_mk, ilo, max_sks_prodi)
        else:
            if is_semester_included:
                list_semester_obj = []
                for semester_prodi_obj, semester_nama in list_filter:
                    list_semester_obj.append(semester_prodi_obj)
                
                # Get mahasiswa's mata kuliah participation
                list_mahasiswa_as_peserta_mk = list_peserta_mk.filter(
                    kelas_mk_semester__mk_semester__semester__in=list_semester_obj,
                    kelas_mk_semester__mk_semester__clo__piclo__performance_indicator__pi_area__ilo=ilo,
                    nilai_akhir__isnull=False,
                ).distinct()

                generate_all_bobot_perhitungan_cpl_mahasiswa(current_worksheet, list_mahasiswa_as_peserta_mk, ilo, max_sks_prodi)
            else:
                list_tahun_ajaran_obj = []
                for tahun_ajaran_prodi_obj, tahun_ajaran_nama in list_filter:
                    list_tahun_ajaran_obj.append(tahun_ajaran_prodi_obj)
                
                # Get mahasiswa's mata kuliah participation
                list_mahasiswa_as_peserta_mk = list_peserta_mk.filter(
                    kelas_mk_semester__mk_semester__semester__tahun_ajaran_prodi__in=list_tahun_ajaran_obj,
                    kelas_mk_semester__mk_semester__clo__piclo__performance_indicator__pi_area__ilo=ilo,
                    nilai_akhir__isnull=False,
                ).distinct()

                generate_all_bobot_perhitungan_cpl_mahasiswa(current_worksheet, list_mahasiswa_as_peserta_mk, ilo, max_sks_prodi)

    # Save the workbook to a BytesIO object
    file_stream = BytesIO()
    workbook.save(file_stream)

    file_stream.seek(0)

    return file_stream


def calculate_pi_elektro_way(
        list_pi: QuerySet[PerformanceIndicator],
        list_mk_semester: QuerySet[MataKuliahSemester],
):
    is_success = False
    message = ''
    result = {}

    # Loop through PI
    for pi in list_pi:
        list_mk_semester_pi = list_mk_semester.filter(
            clo__piclo__performance_indicator=pi
        ).distinct()

        list_nilai_cpmk = []

        # Loop through MK Semester
        for mk_semester in list_mk_semester_pi:
            # Check status nilai MK Semester
            if not mk_semester.status_nilai:
                peserta_qs = PesertaMataKuliah.objects.filter(kelas_mk_semester__mk_semester=mk_semester).exclude(nilai_akhir=None)
                # Skip mata kuliah that doesn't have nilai akhir yet
                if peserta_qs.count() == 0:
                    continue
                
                message = 'Harap melengkapi nilai dari mata kuliah ({}, {}) terlebih dahulu'.format(
                    mk_semester.mk_kurikulum.nama, mk_semester.semester.semester)
                return (is_success, message, result)

            # Query CPMK that has relationship with PI
            cpmk_pi_qs = Clo.objects.filter(
                mk_semester=mk_semester,
                piclo__performance_indicator=pi
            ).distinct()

            # Query Nilai CPMK
            nilai_cpmk_qs = NilaiCloMataKuliahSemester.objects.filter(
                mk_semester=mk_semester,
                clo__in=cpmk_pi_qs
            )

            # Get nilai CPMK
            list_nilai_cpmk.append([
                nilai_cpmk.nilai for nilai_cpmk in nilai_cpmk_qs
            ])

        list_avg_nilai_cpmk = []
        # Calculate average cpmk for each mk
        for nilai_cpmk in list_nilai_cpmk:
            list_avg_nilai_cpmk.append(np.average(nilai_cpmk))

        result[pi.pk] = {
            'pi': pi.deskripsi,
        }

        if len(list_avg_nilai_cpmk) == 0:
            result[pi.pk].update({
                'nilai' : None
            })
        else:
            result[pi.pk].update({
                'nilai' : np.average(list_avg_nilai_cpmk)
            })

    is_success = True
    message = 'Berhasil menghitung capaian PI Program Studi.'

    return (is_success, message, result)


def generate_laporan_pi_pdf(
        list_pi: QuerySet[PerformanceIndicator],
        list_filter,
        calculation_result,
        prodi_name: str, fakultas_name: str,
):
    file_stream = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    page_size = landscape(letter)
    pdf_file = SimpleDocTemplate(file_stream, pagesize=page_size)
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        'Laporan Capaian Pembelajaran Lulusan Mahasiswa',
        style=styles['h1']
    )

    header_style = copy(styles['h2'])
    header_style.alignment = TA_CENTER
    header_content = Paragraph(
        "UNIVERSITAS HASANUDDIN<br/>FAKULTAS {}<br/>PROGRAM STUDI {}".format(fakultas_name, prodi_name), 
        header_style
    )

    # Get the space before and after the paragraph
    space_before = header_content.getSpaceBefore()
    space_after = header_content.getSpaceAfter()

    # Get the height of the paragraph
    para_height = header_content.wrap(pdf_file.width, pdf_file.height)[1]

    # Calculate the total height of the block
    block_height = para_height + space_before + space_after

    frame = Frame(
        pdf_file.leftMargin, 
        pdf_file.bottomMargin, 
        pdf_file.width, 
        pdf_file.height - block_height
    )
    template = PageTemplate(
        frames=frame, 
        onPage=partial(
            export_pdf_header, 
            content=header_content,
            image_path='./static/public/img/logo-unhas.jpg',
            image_width=0.6*inch,
            image_height=0.75*inch,
        )
    )
    pdf_file.addPageTemplates([template])

    # Tabel detail pi
    table_detail_pi_title = Paragraph(
        'Tabel Pencapaian Performance Indicator',
        style=styles['h2']
    )

    # (COL, ROW)
    table_detail_ilo_style_data = [
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),               # Header align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Header font
        ('FONTSIZE', (0, 0), (-1, 1), 11),                  # Header font size
        ('TOPPADDING', (0, 0), (-1, 1), 12),                # Header top padding
        ('BOTTOMPADDING', (0, 0), (-1, 1), 12),             # Header bottom padding
        ('TOPPADDING', (0, 1), (-1, -1), 4),                # Body top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),             # Body bottom padding
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),     # Grid
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),                # Body vertical align
        ('VALIGN', (0, 0), (2, 0), 'MIDDLE'),               # Header vertical align
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),               # Nomor align
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),               # Nilai align
    ]

    list_table_detail_ilo = []
    
    for filter in list_filter:
        table_data = [['No.', 'PI']]
        current_filter_name = filter[1]

        # Add filter name to header
        table_data[0].append(
            Paragraph(
                current_filter_name,
                style=ParagraphStyle(
                    name='table_filter_header',
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold',
                    fontSize=11
                )
            )
        )

        current_table_detail_ilo_style_data = copy(table_detail_ilo_style_data)

        current_table_detail_ilo_style_data += [
            ('SPAN', (2, 0), (-1, 0)),           # Merge filter name cells
        ]

        for i, pi in enumerate(list_pi):
            pi_row = [
                '{}.'.format(i+1),
                Paragraph(pi.deskripsi),
            ]

            current_nilai = calculation_result[current_filter_name][pi.pk]['nilai']
            if current_nilai is None:
                pi_row.append('-')
            else:
                pi_row.append(float('{:.2f}'.format(current_nilai)))


            # Add row to table data
            table_data.append(pi_row)
        
        current_table_detail_ilo_style = TableStyle(current_table_detail_ilo_style_data)
        # Make table
        detail_ilo_table = Table(
            table_data, 
            style=current_table_detail_ilo_style,
            hAlign='LEFT',
            colWidths=[1*cm, None, None],
            repeatRows=1,
        )

        list_table_detail_ilo.append(detail_ilo_table)

    # Build
    pdf_file_elements = [
        title,
        table_detail_pi_title,
    ]
    for detail_ilo_table in list_table_detail_ilo:
        pdf_file_elements += [detail_ilo_table, PageBreak()]
    
    pdf_file.build(pdf_file_elements)
    
    # Close the PDF object cleanly
    file_stream.seek(0)

    return file_stream
